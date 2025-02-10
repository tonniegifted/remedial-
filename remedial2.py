from cProfile import label
from tkinter import*
from tkinter import ttk
from tkinter import messagebox
# from CTkMessagebox import
from CTkMessagebox import CTkMessagebox
import mysql.connector
from customtkinter import CTkToplevel
from openpyxl import Workbook
from openpyxl.styles import Font
from decimal import Decimal
import customtkinter as ctk
from datetime import date
import ttkbootstrap as tb
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side
from tkinter import filedialog, Tk
from datetime import datetime
import os
#connecting to database
my_db=mysql.connector.connect(host="localhost",
                              user="root",
                              password="print()",
                              database="remedial")

# creating cursor
cur=my_db.cursor()

#creating connection for sqlite3    "C:/Users/Administrator/Documents"
#"C:/Users/Administrator/Documents/Remedial App/Remedial App.db"

path="C:/Users/Administrator/Documents/Remedial App/Remedial App.db"
# my_db = sqlite3.connect(path)
#creating cursor
cur=my_db.cursor()
root= ctk.CTk()
# ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")
width=root.winfo_width()
height=root.winfo_height()
root.title("Remedial App")
path="C:/Users/Administrator/Documents/Remedial App/Resources/remedial convert.ico"
root.iconbitmap(path)
# root.geometry(f"{width}x{height}+0+0")
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
# Set the window size
window_width = 680
window_height = 600
# Calculate position to center horizontally and start at the top
position_x = (screen_width - window_width) // 2
position_y = 10  # Start at the to
root.geometry(f"{window_width}x{window_height}+{position_x}+{position_y}")
root.resizable(width=False, height=False)
#color blue
blue="#4582EC"



#FUNCTIONS
def generate_records(e=None):
    generate_teacher_attendance_report()
    generate_grade_reports()
def generate_teacher_attendance_report():
    try:
        # Fetch school name
        school_name_path = "C:/Users/Administrator/Documents/Remedial2/name.txt"
        with open(school_name_path, "r") as file:
            school_name = file.read().strip()

        # Default save path
        default_folder = "C:/Users/Administrator/Documents/Remedial2"
        os.makedirs(default_folder, exist_ok=True)

        # Ask user for folder selection, fallback to default
        folder_path = filedialog.askdirectory(title="Select Save Location") or default_folder

        file_path = os.path.join(folder_path, "Teacher_Attendance_Report.xlsx")

        # Check if file exists
        if os.path.exists(file_path):
            replace = messagebox.askyesno(
                "File Exists",
                f"{file_path} already exists. Do you want to replace it?"
            )
            if not replace:
                return  # Exit if the user doesn't want to replace the file

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Teacher Attendance Report"

        # === Insert Logo at A2 ===
        logo_path = "C:/Users/Administrator/Documents/Remedial2/LOGO FINALE.png"
        try:
            img = Image(logo_path)
            img.anchor = "A2"  # Position logo
            ws.add_image(img)
        except Exception as e:
            messagebox.showwarning("Warning", f"Could not load logo: {e}")

        # === Add School Details at C2 ===
        ws.merge_cells("C2:H2")
        ws["C2"] = school_name
        ws["C2"].font = Font(size=14, bold=True)
        ws["C2"].alignment = Alignment(horizontal="center", vertical="center")

        ws["C3"] = f"Date Generated: {datetime.today().strftime('%d-%m-%Y')}"
        ws["C3"].font = Font(size=12, bold=True)
        ws["C3"].alignment = Alignment(horizontal="left")

        # === Table Headers ===
        headers = ["S/No", "Name", "Grade", "Subject", "Token Amount", "Session", "Week", "Date"]
        start_row = 9
        start_col = 2

        for col_index, header in enumerate(headers, start=start_col):
            cell = ws.cell(row=start_row, column=col_index, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=Side(style="thin"))

        # === Fetch Data from MySQL ===
        cur.execute("""
            SELECT t.first, t.second, a.grade, a.session,
                   a.record_date, w.selected_week, a.session_amount, a.subject
            FROM teacher_attendance a
            JOIN teacher t ON t.teacher_id = a.teacher_id
            JOIN week_number w ON w.week_number_id = a.week_number_id
            JOIN term tm ON a.term_id = tm.term_id
            WHERE a.term_id = (SELECT term_id FROM term WHERE is_active = 1)
        """)
        items = cur.fetchall()

        data_start_row = start_row + 1
        if items:
            for row_index, item in enumerate(items, start=1):
                full_name = f"{item[0]} {item[1]}".title()
                formatted_date = item[4].strftime("%d-%m-%Y")  # Format date

                # Insert data into the worksheet
                ws.cell(row=data_start_row + row_index - 1, column=start_col, value=row_index)  # Serial Number
                ws.cell(row=data_start_row + row_index - 1, column=start_col + 1, value=full_name)  # Name
                ws.cell(row=data_start_row + row_index - 1, column=start_col + 2, value=item[2])  # Grade
                ws.cell(row=data_start_row + row_index - 1, column=start_col + 3, value=item[7])  # Subject
                ws.cell(row=data_start_row + row_index - 1, column=start_col + 4, value=item[6])  # Token Amount
                ws.cell(row=data_start_row + row_index - 1, column=start_col + 5, value=item[3])  # Session
                ws.cell(row=data_start_row + row_index - 1, column=start_col + 6, value=item[5])  # Week
                ws.cell(row=data_start_row + row_index - 1, column=start_col + 7, value=formatted_date)  # Date
        else:
            ws.cell(row=data_start_row, column=start_col, value="No records found.")

        # === Adjust Column Width ===
        ws.column_dimensions['C'].width = 15  # Name
        ws.column_dimensions['D'].width = 10  # Grade
        ws.column_dimensions['E'].width = 12  # Subject
        ws.column_dimensions['F'].width = 12  # Token Amount
        ws.column_dimensions['G'].width = 15  # Session
        ws.column_dimensions['H'].width = 15  # Week
        ws.column_dimensions['I'].width = 15  # Date

        # === Ensure Grid Lines Are Printable ===
        ws.print_options.gridLines = True

        # === A4 Page Setup ===
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        # Save the file
        wb.save(file_path)

        messagebox.showinfo("Success", "Teacher Attendance Report Generated Successfully!")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")




def generate_grade_reports():
    try:
        # Fetch school name
        school_name_path = "C:/Users/Administrator/Documents/Remedial2/name.txt"
        with open(school_name_path, "r") as file:
            school_name = file.read().strip()

        # Fetch active term
        cur.execute("SELECT selected_term FROM term WHERE is_active=1")
        term_data = cur.fetchone()
        if not term_data:
            messagebox.showerror("Error", "No active term found in the database.")
            return
        
        term = term_data[0]

        # Default save path
        default_folder = "C:/Users/Administrator/Documents/Remedial2"
        os.makedirs(default_folder, exist_ok=True)

        # Ask user for folder selection, fallback to default
        folder_path = filedialog.askdirectory(title="Select Save Location") or default_folder

        grades = ["Seven", "Eight", "Nine"]

        for grade in grades:
            file_path = os.path.join(folder_path, f"Grade_{grade}_Balance.xlsx")

            # Check if file exists
            if os.path.exists(file_path):
                replace = messagebox.askyesno(
                    "File Exists",
                    f"{file_path} already exists. Do you want to replace it?"
                )
                if not replace:
                    continue  # Skip this grade and move to the next

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Grade {grade} Report"

            # === Insert Logo at A2 ===
            logo_path = "C:/Users/Administrator/Documents/Remedial2/LOGO FINALE.png"
            try:
                img = Image(logo_path)
                img.anchor = "A2"  # Position logo
                ws.add_image(img)
            except Exception as e:
                messagebox.showwarning("Warning", f"Could not load logo for Grade {grade}: {e}")

            # === Fetch Term Fee for the Grade from `termpay` ===
            cur.execute("SELECT lnr_pay FROM termly_pay")
            term_fee_result = cur.fetchone()
            term_fee = term_fee_result[0] if term_fee_result else 0

            # === Add School Details at C2 ===
            ws.merge_cells("C2:H2")
            ws["C2"] = school_name
            ws["C2"].font = Font(size=14, bold=True)
            ws["C2"].alignment = Alignment(horizontal="center", vertical="center")

            ws["C3"] = f"Grade: {grade}"
            ws["C4"] = f"{term}"
            ws["C6"] = f"{datetime.today().strftime('%d-%m-%Y')}"

            for cell in ["C3", "C4", "C6"]:
                ws[cell].font = Font(size=12, bold=True)
                ws[cell].alignment = Alignment(horizontal="right")

            # === Table Headers ===
            headers = ["#.", "Adm", "Name", "Paid", "Bal"]
            start_row = 9
            start_col = 2

            for col_index, header in enumerate(headers, start=start_col):
                cell = ws.cell(row=start_row, column=col_index, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(bottom=Side(style="thin"))

            # === Fetch Data from MySQL and Sort by Amount Paid (Desc) ===
            cur.execute("""
                SELECT 
                    learner.learner_id, 
                    CONCAT(learner.first, ' ', learner.second, ' ', learner.surname) AS full_name, 
                    IFNULL(SUM(transactions.amount_paid), 0) AS amount_paid, 
                    %s - IFNULL(SUM(transactions.amount_paid), 0) AS balance
                FROM learner
                LEFT JOIN transactions 
                    ON learner.learner_id = transactions.learner_id 
                    AND transactions.term_id = (SELECT term_id FROM term WHERE is_active = 1)
                WHERE learner.grade = %s
                GROUP BY learner.learner_id
                ORDER BY amount_paid DESC, learner.surname ASC
            """, (term_fee, grade))

            learners = cur.fetchall()

            data_start_row = start_row + 1
            if learners:
                for row_index, learner in enumerate(learners, start=1):
                    ws.cell(row=data_start_row + row_index - 1, column=start_col, value=row_index)
                    for col_index, value in enumerate(learner, start=start_col + 1):
                        ws.cell(row=data_start_row + row_index - 1, column=col_index, value=value)
            else:
                ws.cell(row=data_start_row, column=start_col, value="No records found.")

            # === Adjust Column Width ===
            ws.column_dimensions['C'].width = 12  # Learner ID
            ws.column_dimensions['D'].width = 25  # Full Name
            ws.column_dimensions['E'].width = 15  # Amount Paid
            ws.column_dimensions['F'].width = 15  # Balance

            # === Ensure Grid Lines Are Printable ===
            ws.print_options.gridLines = True

            # === A4 Page Setup ===
            ws.page_setup.paperSize = ws.PAPERSIZE_A4

            # Save the file
            wb.save(file_path)

        messagebox.showinfo("Success", "All Grade Reports Generated Successfully!")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")



def save_school_name():
    global school_name_entry,school_name_disp_label
    path="c:/Users/Administrator/Documents/Remedial2/name.txt"
    name=school_name_entry.get().title().strip()
    if name:
        with open(f"{path}","w") as file:
            file.write(f"{name}")
            messagebox.showinfo("School Name",f"School Name '{name.title()}'\nsaved successfully")
            school_name_entry.delete(0,END)
            
    else:
        root.bell()
        messagebox.showwarning("School Name","School Name cannot be blank")
def school_name_func():
    global school_name_entry,school_name_disp_label
    school_name_win = CTkToplevel(root)
    school_name_win.geometry("400x150+390+10")
    # style
    school_name_win.resizable(width=False,height=False)
    school_name_win.transient(root)
    school_name_win.grab_set()
    school_name_win.title("School_name")
    school_name_win.after(100, lambda: school_name_win.lift())  # Delay lifting the window
    school_name_win.after(200, lambda: school_name_win.focus_force())  # Delay forcing
    # lr_archive_frame=ctk.CTkFrame(school_name_win,fg_color="transparent")
    widget_frame=ctk.CTkFrame(school_name_win,fg_color="transparent")
    widget_frame.place(x=20,y=20)
    school_name_label=ctk.CTkLabel(widget_frame,text="School Name",font=("helvetica",16))
    school_name_label.grid(row=0,column=0,sticky=W)
    school_name_entry=ctk.CTkEntry(widget_frame,font=("helvetica",16),width=250,border_color=blue)
    school_name_entry.grid(row=0,column=1,padx=5)
    school_name_button=ctk.CTkButton(widget_frame,text="Save",width=80,command=save_school_name)
    school_name_button.grid(row=1,columnspan=2,pady=10)
    school_name_disp_label=ctk.CTkLabel(widget_frame,text="",font=("helvetica",16))
    school_name_disp_label.grid(row=2,column=0)
def delete_archive():
    global archive_label
    resp=messagebox.askyesno("Archives","Are you sure you want to delete learner\narchive records?")
    if resp:
        cur.execute("DELETE FROM archive")
        my_db.commit()
        disp_label.configure("Remedial App","Learner Archive records Deleted\nsuccessfully")
        # archive_function()
        # display_archives()
    else:
        pass
#
def display_archives():
    global lr_archive_win,archive_window
    cur.execute("SELECT * FROM archive")
    item=cur.fetchall()
    if item:
        lr_archive_win = CTkToplevel(root)
        lr_archive_win.geometry("530x380+390+10")
        # style
        lr_archive_win.resizable(width=False,height=False)
        lr_archive_win.transient(root)
        lr_archive_win.grab_set()
        lr_archive_win.title("Archived Learners")
        lr_archive_win.after(100, lambda: lr_archive_win.lift())  # Delay lifting the window
        lr_archive_win.after(200, lambda: lr_archive_win.focus_force())  # Delay forcing
        lr_archive_frame=ctk.CTkFrame(lr_archive_win,fg_color="transparent")
        archive_tree_scroll=Scrollbar(lr_archive_frame,orient=VERTICAL)
        archive_tree_scroll.pack(side=RIGHT,fill=Y)
        lr_archive_frame.place(x=20,y=30)
        lr_archive_tree=ttk.Treeview(lr_archive_frame,yscrollcommand=archive_tree_scroll.set,height=12)
        lr_archive_tree.pack()
        archive_tree_scroll.configure(command=lr_archive_tree.yview)
        #defining columns
        lr_archive_tree["columns"]=('s/no','name','archived')
        lr_archive_tree.column("#0",width=0,stretch=NO)
        lr_archive_tree.column("s/no",width=40,anchor="center",minwidth=35)
        lr_archive_tree.column("name",width=270,minwidth=180,anchor=W)
        lr_archive_tree.column("archived",width=180,minwidth=180,anchor=W)
        #headings
        lr_archive_tree.heading("#0",text="")
        lr_archive_tree.heading("s/no",text="#",anchor=CENTER)
        lr_archive_tree.heading("name",text="NAME",anchor=CENTER)
        lr_archive_tree.heading("archived",text="DATE ARCHIVED",anchor=CENTER)
        # defining columns
        # fee collection treeview
        # lr_archive_tree = ctk.CTkFrame(lr_archive_win, fg_color="transparent")
        widget_frame=ctk.CTkFrame(lr_archive_win,fg_color="transparent")
        widget_frame.place(x=20,y=320)
        archive_del_button=ctk.CTkButton(widget_frame,text="Delete",command=delete_archive)
        archive_del_button.grid(row=0,column=0,sticky=W,padx=5)
        archive_label=ctk.CTkLabel(widget_frame,text="")
        archive_label.grid(row=0,column=1,)
        for index,item in enumerate(item,start=1):
            time=item[5]
            formatted_time=time.strftime("%d/%m/%Y")
            

            full_name=f"{item[1]} {item[2]} {item[3]}".title()
            lr_archive_tree.insert("",END,values=(index,full_name,formatted_time))
    
    else:
        messagebox.showinfo("Archives","No records in Archives")
       

def delete_attend_record():
    global tr_attend_tree,tr_attend_win
    try:
        res=messagebox.askyesno("Teacher Attendance History","Are you sure you want to delete\nThe last entered attendance record?")
        if res:
            # pos=tr_attend_tree.selection()
            # value=tr_attend_tree.item(pos,"values")
            # teacher_id=value[0]
            cur.execute("""SELECT teacher_attendance_id FROM teacher_attendance ORDER BY
                        record_date DESC LIMIT 1""")
            tr_attend_id=cur.fetchone()[0]
            # deleting last entered record by tr_id incase of a mistake
        cur.execute("""DELETE FROM teacher_attendance WHERE teacher_attendance_id=%s""",
                    (tr_attend_id,))
        my_db.commit()
    # tr_attend_tree.delete(value)
        messagebox.showinfo("Teacher Attendance History","Recent teacher attendance record deleted\nsuccessfully")
    except:
        messagebox.showerror("Teacher Attendance History","Unexpected error just occurred")
    # cur.execute("DELETE FROM teacher_attendance")
    # finally:
    #     tr_attend_win.destroy()
#delete teacher attendant archive
def delete_teacher_attendance():
    global tr_attend_archive_tree, tr_attend_archive_search_entry, tr_attend_archive_disp_label, tr_attend_archive_win
    selected_records = tr_attend_archive_tree.selection()

    if not selected_records:
        messagebox.showwarning("Remedial App", "Select records before attempting to delete.")
        return

    root.bell()
    resp = messagebox.askyesno("Remedial App",
                               "Are you sure you want to delete the selected TEACHER ATTENDANCE records permanently?")

    if resp:
        teacher_ids = []
        for item in selected_records:
            values = tr_attend_archive_tree.item(item, "values")
            teacher_ids.append((values[1],))  # Assuming teacher_id is at index 1

            # Delete from treeview
            tr_attend_archive_tree.delete(item)

        # Execute delete operation
        cur.executemany("DELETE FROM teacher_attendance_archive WHERE teacher_id = %s", teacher_ids)
        my_db.commit()


        # Execute delete operation if teacher IDs exist
        if teacher_ids:
            cur.executemany("DELETE FROM teacher_attendance_archive WHERE teacher_id = %s", teacher_ids)
            my_db.commit()
            messagebox.showinfo("Remedial App", "Records deleted successfully.")
        tr_attend_archive_win.destroy()
        disp_attend_archive_func()

def disp_attend_archive_func(e=None):
    global tr_attend_archive_tree, tr_attend_archive_search_entry, tr_attend_archive_disp_label, tr_attend_archive_win
    cur.execute("""SELECT t.teacher_id, t.title, t.first,t.second,t.surname, ar.grade,ar.session,
                ar.record_date,w.selected_week,ar.session_amount,ar.subject
                FROM teacher_attendance_archive ar JOIN teacher t ON t.teacher_id=ar.teacher_id
                JOIN week_number w ON w.week_number_id=ar.week_number_id JOIN term tm
                ON ar.term_id=tm.term_id
                WHERE  ar.term_id=(SELECT term_id FROM term WHERE is_active=1)""")
    # ID-0,name-(1-4),grade-5,subject-10,token-9,session-6,week-8,date-7
    items = cur.fetchall()
    if items:
        tr_attend_archive_win = CTkToplevel(root)
        tr_attend_archive_win.geometry("870x400+390+10")
        tr_attend_archive_win.title("Teacher Attendance History")
        # style
        tr_attend_archive_win.resizable(width=False,height=False)
        tr_attend_archive_win.transient(root)
        tr_attend_archive_win.grab_set()
        tr_attend_archive_win.title("Tr. Attendance Archive")
        tr_attend_archive_win.after(100, lambda: tr_attend_archive_win.lift())  # Delay lifting the window
        tr_attend_archive_win.after(200, lambda: tr_attend_archive_win.focus_force())  # Delay forcing
        # defining columns
        # fee collection treeview
        tr_attend_archive_frame = ctk.CTkFrame(tr_attend_archive_win, fg_color="transparent")
        tr_attend_archive_tree_scroll = Scrollbar(tr_attend_archive_frame, orient=VERTICAL)
        tr_attend_archive_tree_scroll.pack(side=RIGHT, fill=Y)
        tr_attend_archive_frame.place(x=20, y=50)
        tr_attend_archive_tree = ttk.Treeview(tr_attend_archive_frame, yscrollcommand=tr_attend_archive_tree_scroll.set, height=12)
        tr_attend_archive_tree_scroll.configure(command=tr_attend_archive_tree.yview)
        tr_attend_archive_tree["columns"] = ('s/no','ID', 'name', 'grade', 'subject', 'token', 'session', 'week', 'date')
        tr_attend_archive_tree.column("#0", width=0, stretch=NO)
        tr_attend_archive_tree.column("s/no", width=45, anchor="center", minwidth=30)
        tr_attend_archive_tree.column("ID", width=45, anchor="center", minwidth=30)
        tr_attend_archive_tree.column("name", width=150, minwidth=150, anchor=W)
        tr_attend_archive_tree.column("grade", width=80, minwidth=75, anchor=CENTER)
        tr_attend_archive_tree.column("subject", width=60, minwidth=50, anchor=CENTER)
        tr_attend_archive_tree.column("token", width=70, minwidth=50, anchor=CENTER)
        tr_attend_archive_tree.column("session", width=100, minwidth=90, anchor=W)
        tr_attend_archive_tree.column("week", width=150, minwidth=150, anchor=CENTER)
        tr_attend_archive_tree.column("date", width=120, minwidth=120, anchor=W)
        # headings
        tr_attend_archive_tree.heading("#0", text="")
        tr_attend_archive_tree.heading("s/no", text="#", anchor=CENTER)
        tr_attend_archive_tree.heading("ID", text="ID", anchor=CENTER)
        tr_attend_archive_tree.heading("name", text="NAME", anchor=CENTER)
        tr_attend_archive_tree.heading("grade", text="GRADE", anchor=CENTER)
        tr_attend_archive_tree.heading("subject", text="SUBJ", anchor=CENTER)
        tr_attend_archive_tree.heading("token", text="TOKEN", anchor=CENTER)
        tr_attend_archive_tree.heading("session", text="SESSION", anchor=CENTER)
        tr_attend_archive_tree.heading("week", text="WEEK", anchor=CENTER)
        tr_attend_archive_tree.heading("date", text="DATE", anchor=CENTER)
        # fee collection frame display
        tr_attend_archive_tree.pack(fill=BOTH, expand=True)
        # widgets frame
        widgets_frame = ctk.CTkFrame(tr_attend_archive_win, fg_color="transparent")
        widgets_frame.place(x=20, y=10)
        # search2
        # tr_attend_archive_search_button=Button(tr_attend_archive_win,text="Search",font="times 11",command=lambda: disp_attendance_history(None))
        tr_attend_archive_search_button = ctk.CTkButton(widgets_frame, text="Search", command=search_teacher_attend_archive, width=10)
        tr_attend_archive_search_button.grid(row=0, column=1)
        tr_attend_archive_search_entry = ctk.CTkEntry(widgets_frame, width=80, border_color=blue, placeholder_text="Tr ID")

        # tr_attend_archive_search_entry.place(x=100, y=10)
        tr_attend_archive_search_entry.grid(row=0, column=0, padx=5)
        tr_attend_archive_disp_label = ctk.CTkLabel(widgets_frame, font=("Helvetica", 12), text="")
        # tr_attend_archive_disp_label.place(x=250, y=10)
        tr_attend_archive_disp_label.grid(row=0, column=3, padx=10)
        tr_archive_delete_button=ctk.CTkButton(tr_attend_archive_win,text="Delete",command=delete_teacher_attendance,width=80)
        tr_archive_delete_button.place(x=30,y=330)

        for index, items in enumerate(items, start=1):
            time = items[7]
            formatted_time = time.strftime("%d-%m-%Y")
            # ID-0,name-(1-4),grade-5,subject-10,token-9,session-6,week-8,date-7

            full_name = f"{items[1]} {items[2]} {items[3]}".title()
            tr_attend_archive_tree.insert("", END, values=(
                index,items[0] ,full_name, items[5], items[10], items[9], items[6], items[8], formatted_time))
    else:
        messagebox.showinfo("Remedial App","No teacher attendance record found")
def search_teacher_attend_archive():
    global tr_attend_archive_tree, tr_attend_archive_search_entry, tr_attend_archive_disp_label, tr_attend_archive_win
    tr_attend_archive_tree.delete(*tr_attend_archive_tree.get_children())
    try:
        teacher_id = tr_attend_archive_search_entry.get()
        teacher_id = int(teacher_id)
        if not teacher_id:
            tr_attend_archive_disp_label.configure(text="Search box cannot be blank")
            tr_attend_archive_disp_label.after(4000, lambda: tr_attend_archive_disp_label.configure(text=""))
            return
        # displaying teacher attendance
        cur.execute("""SELECT t.first,t.second, ar.grade,ar.session,
            ar.record_date,w.selected_week,ar.session_amount,ar.subject
            FROM teacher_attendance_archive ar JOIN teacher t ON t.teacher_id=ar.teacher_id
            JOIN week_number w ON w.week_number_id=ar.week_number_id JOIN term tm
            ON ar.term_id=tm.term_id
            WHERE ar.teacher_id=%s AND ar.term_id=(SELECT term_id FROM term WHERE is_active=1)""", (teacher_id,))
        items = cur.fetchall()
        if items:
            for index, items in enumerate(items, start=1):
                time = items[4]
                formatted_time = time.strftime("%d-%m-%Y")

                full_name = f"{items[0]} {items[1]}".title()
                tr_attend_archive_tree.insert("", END, values=(
                index, full_name, items[2], items[7], items[6], items[3], items[5], formatted_time))
            cur.execute("""SELECT SUM(session_amount) FROM teacher_attendance_archive
                        WHERE teacher_id=%s""", (teacher_id,))
            am = cur.fetchone()
            if am:
                amount_paid = am[0]
                tr_attend_archive_disp_label.configure(text=f"Total Paid= {amount_paid}", font=("times",14 ))
                tr_attend_archive_disp_label.after(4000,lambda:tr_attend_archive_disp_label.configure(text=""))

        else:
            tr_attend_archive_disp_label.configure(text=f"Record with Tr.No {tr_attend_archive_search_entry.get()} not found")
            tr_attend_archive_disp_label.after(4000, lambda: tr_attend_archive_disp_label.configure(text=""))
            tr_attend_archive_search_entry.delete(0, END)
    except Exception as ex:
        # tr_attend_archive_disp_label.configure(text=f"{ex}")
        # tr_attend_archive_disp_label.after(4000, lambda: tr_attend_archive_disp_label.configure(text=""))
        messagebox.showerror("Remedial App",f"{ex}")
#sets the treeview for teacher attendance history
def tr_attendance_func(e=None):
    global tr_attend_tree, tr_attend_search_entry, tr_attend_disp_label, tr_attend_win
    cur.execute("""SELECT t.first,t.second, a.grade,a.session,
               a.record_date,w.selected_week,a.session_amount,a.subject
               FROM teacher_attendance a JOIN teacher t ON t.teacher_id=a.teacher_id
               JOIN week_number w ON w.week_number_id=a.week_number_id JOIN term tm
               ON a.term_id=tm.term_id
               WHERE a.term_id=(SELECT term_id FROM term WHERE is_active=1)""")
    items = cur.fetchall()
    if items:
        tr_attend_win = CTkToplevel(root)
        tr_attend_win.geometry("810x400+390+10")
        tr_attend_win.title("Teacher Attendance History")
        # style
        # tr_attend_win.resizable(width=False,height=False)
        tr_attend_win.transient(root)
        tr_attend_win.grab_set()
        tr_attend_win.title("Monitor Tr. Attendance")
        tr_attend_win.after(100, lambda: tr_attend_tree.lift())  # Delay lifting the window
        tr_attend_win.after(200, lambda: tr_attend_tree.focus_force())  # Delay forcing
        # defining columns
        # fee collection treeview
        tr_attend_frame = ctk.CTkFrame(tr_attend_win, fg_color="transparent")
        tr_attend_tree_scroll = Scrollbar(tr_attend_frame, orient=VERTICAL)
        tr_attend_tree_scroll.pack(side=RIGHT, fill=Y)
        tr_attend_frame.place(x=20, y=50)
        tr_attend_tree = ttk.Treeview(tr_attend_frame, yscrollcommand=tr_attend_tree_scroll.set, height=12)
        tr_attend_tree_scroll.configure(command=tr_attend_tree.yview)
        tr_attend_tree["columns"] = ('s/no', 'name', 'grade', 'subject', 'token', 'session', 'week', 'date')
        tr_attend_tree.column("#0", width=0, stretch=NO)
        tr_attend_tree.column("s/no", width=40, anchor="center", minwidth=35)
        tr_attend_tree.column("name", width=150, minwidth=150, anchor=W)
        tr_attend_tree.column("grade", width=80, minwidth=75, anchor=CENTER)
        tr_attend_tree.column("subject", width=60, minwidth=50, anchor=CENTER)
        tr_attend_tree.column("token", width=70, minwidth=50, anchor=CENTER)
        tr_attend_tree.column("session", width=100, minwidth=90, anchor=W)
        tr_attend_tree.column("week", width=150, minwidth=150, anchor=CENTER)
        tr_attend_tree.column("date", width=120, minwidth=120, anchor=W)
        # headings
        tr_attend_tree.heading("#0", text="")
        tr_attend_tree.heading("s/no", text="#", anchor=CENTER)
        tr_attend_tree.heading("name", text="NAME", anchor=CENTER)
        tr_attend_tree.heading("grade", text="GRADE", anchor=CENTER)
        tr_attend_tree.heading("subject", text="SUBJ", anchor=CENTER)
        tr_attend_tree.heading("token", text="TOKEN", anchor=CENTER)
        tr_attend_tree.heading("session", text="SESSION", anchor=CENTER)
        tr_attend_tree.heading("week", text="WEEK", anchor=CENTER)
        tr_attend_tree.heading("date", text="DATE", anchor=CENTER)
        # fee collection frame display
        tr_attend_tree.pack(fill=BOTH, expand=True)
        # widgets frame
        widgets_frame=ctk.CTkFrame(tr_attend_win,fg_color="transparent")
        widgets_frame.place(x=20,y=10)
        # search2
        # tr_attend_search_button=Button(tr_attend_win,text="Search",font="times 11",command=lambda: disp_attendance_history(None))
        tr_attend_search_button = ctk.CTkButton( widgets_frame, text="Search", command=disp_attendance_history,width=10)
        tr_attend_search_button.grid(row=0,column=1)
        tr_attend_search_entry = ctk.CTkEntry( widgets_frame, width=80,border_color=blue,
                                              placeholder_text="Tr ID")

        # tr_attend_search_entry.place(x=100, y=10)
        tr_attend_search_entry.grid(row=0,column=0,padx=5)
        # button to delete teacher attendance per id
        # tr_attend_delete_button = Button(tr_attend_win, text="Delete", font="times 11", command=delete_attend_record)
        # tr_attend_delete_button.place(x=30, y=380)
        # display what should be searched
        # display_label2
        tr_attend_disp_label = ctk.CTkLabel( widgets_frame, font=("Helvetica",12),text="")
        # tr_attend_disp_label.place(x=250, y=10)
        tr_attend_disp_label.grid(row=0,column=3,padx=10)
        tr_attend_delete_button=ctk.CTkButton(tr_attend_win,text="Delete",
                                              command=delete_attend_record,width=80)
        tr_attend_delete_button.place(x=40,y=330)
        for index, items in enumerate(items, start=1):
            time = items[4]
            formatted_time = time.strftime("%d-%m-%Y")

            full_name = f"{items[0]} {items[1]}".title()
            tr_attend_tree.insert("", END, values=(
            index, full_name, items[2], items[7], items[6], items[3], items[5], formatted_time))

    else:
        messagebox.showinfo("Transaction history", "No Teacher Attendance Records")
        # tr_attend_win.destroy()

#displaying attendance history searches
def disp_attendance_history():
    global tr_attend_tree, session_combo, grade_combo, subject_combo, pay_token_entry, tr_win_disp,tr_attend_search_entry,tr_attend_disp_label

    try:
        # Ensure treeview is cleared before inserting new records
        teacher_id=int(tr_attend_search_entry.get())
        tr_attend_tree.delete(*tr_attend_tree.get_children())
        # Query to get teacher attendance records
        cur.execute("""
            SELECT t.teacher_id,t.title,t.first, t.second, a.grade, a.session, a.record_date, 
                   w.selected_week, a.session_amount, a.subject
            FROM teacher_attendance a 
            JOIN teacher t ON t.teacher_id = a.teacher_id
            JOIN week_number w ON w.week_number_id = a.week_number_id 
            JOIN term tm ON a.term_id = tm.term_id
            WHERE a.teacher_id = %s 
            AND a.term_id = (SELECT term_id FROM term WHERE is_active = 1)
        """, (teacher_id,))

        items = cur.fetchall()
        if items:
            for index, item in enumerate(items, start=1):
                formatted_time = item[4].strftime("%d-%m-%Y")
                full_name = f"{item[0]} {item[1]}".title()
                tr_attend_tree.insert("", "end", values=(
                    index, full_name, item[2], item[7], item[6], item[3], item[5], formatted_time
                ))

            # Fetch the teacher's paid tokens
            cur.execute("SELECT token_paid FROM teacher_token WHERE teacher_id = %s", (teacher_id,))
            tr_paid = cur.fetchone()
            teacher_paid = int(tr_paid[0]) if tr_paid else 0

            # Fetch the total session amount
            cur.execute("SELECT SUM(session_amount) FROM teacher_attendance WHERE teacher_id = %s", (teacher_id,))
            tot = cur.fetchone()
            total = int(tot[0]) if tot and tot[0] else 0

            balance = total - teacher_paid

            # Ensure `tr_win_disp` exists before modifying it
            if tr_attend_disp_label and tr_attend_disp_label.winfo_exists():
                tr_attend_disp_label.configure(text=f"Balance = {balance}")
                tr_attend_disp_label.after(4000, lambda: tr_attend_disp_label.configure(text=""))
            # else:
            #     print("Warning: tr_win_disp is closed or does not exist.")
        else:
            tr_attend_disp_label.configure(text=f"Record with Tr.No {tr_attend_search_entry.get()} not found")
            tr_attend_disp_label.after(4000, lambda: tr_attend_disp_label.configure(text=""))

        # Clear search entry
        tr_attend_search_entry.delete(0, "end")

    except Exception as e:
        tr_attend_disp_label.configure(text="An error occurred. Please try again.")
        tr_attend_disp_label.after(4000, lambda: tr_attend_disp_label.configure(text=""))

def pay_token():
    global teacher_tree, session_combo, grade_combo, subject_combo, pay_token_entry, tr_win_disp
    try:
        pos=teacher_tree.selection()
        value = teacher_tree.item(pos,"values")
        teacher_id = value[1]

        # Get the active term ID
        cur.execute("SELECT term_id FROM term WHERE is_active=1")
        term_id = cur.fetchone()
        if term_id:
            active_term_id = term_id[0]
        else:
            messagebox.showerror("Remedial App", "No active term found.")
            return

        # Check if the teacher has any attendance records
        cur.execute("""
            SELECT COUNT(*)
            FROM teacher_attendance
            WHERE teacher_id = %s
        """, (teacher_id,))
        attendance_count = cur.fetchone()[0]

        if attendance_count == 0:
            messagebox.showerror("Remedial App", "Tokens cannot be paid for a teacher without\nattendance records.")
            return

        # Calculate the total session amount for the teacher's attendance records
        cur.execute("""
            SELECT SUM(session_amount)
            FROM teacher_attendance
            WHERE teacher_id = %s
        """, (teacher_id,))
        total_session_amount = cur.fetchone()[0]
        total_session_amount = Decimal(total_session_amount) if total_session_amount else Decimal(0)

        # Get the token amount paid
        token_paid = pay_token_entry.get()
        try:
            token_paid = Decimal(token_paid)
        except ValueError:
            messagebox.showerror("Remedial App", "Invalid token amount. Please enter a valid number.")
            return

        # Validate that the token paid does not exceed the total session amount
        if token_paid > total_session_amount:
            messagebox.showerror("Remedial App",
                                 f"Token payment exceeds the total\nsession amount ({total_session_amount}).")
            return

        # Check the current token balance in the teacher_token table
        cur.execute("""
            SELECT token_paid 
            FROM teacher_token 
            WHERE teacher_id = %s AND term_id = %s
        """, (teacher_id, active_term_id))
        token = cur.fetchone()

        current_token_balance = Decimal(token[0]) if token else Decimal(0)

        # Ensure the entered token amount does not exceed the available balance
        required_balance = total_session_amount - current_token_balance
        if token_paid > required_balance:
            messagebox.showerror("Remedial App",
                                 f"The entered token amount exceeds the \nrequired balance of {required_balance}.")
            return

        # Add the token payment to the current balance
        new_token = current_token_balance + token_paid

        # Insert or update the teacher's token record in the teacher_token table
        cur.execute("""
            INSERT INTO teacher_token (teacher_id, term_id, token_paid)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE token_paid = VALUES(token_paid)
        """, (teacher_id, active_term_id, new_token))
        my_db.commit()

        # Retrieve attendance records in ascending order
        cur.execute("""
            SELECT teacher_attendance_id, session_amount
            FROM teacher_attendance
            WHERE teacher_id = %s
            ORDER BY teacher_attendance_id ASC
        """, (teacher_id,))
        attendance_records = cur.fetchall()

        payment_amount = new_token
        surplus_token = Decimal(0)

        # Process attendance records
        for record in attendance_records:
            attendance_id, session_amount = record
            session_amount = Decimal(session_amount)

            if payment_amount + surplus_token >= session_amount:
                # Archive attendance record
                cur.execute("""
                    INSERT INTO teacher_attendance_archive 
                        (teacher_id, week_number_id, session, grade, session_amount, subject, term_id)
                    SELECT teacher_id, week_number_id, session, grade, session_amount, subject, term_id
                    FROM teacher_attendance
                    WHERE teacher_attendance_id = %s
                """, (attendance_id,))

                # Delete the attendance record
                cur.execute("""
                    DELETE FROM teacher_attendance
                    WHERE teacher_attendance_id = %s
                """, (attendance_id,))

                # Deduct the session amount from payment
                payment_amount -= session_amount
            else:
                surplus_token = payment_amount + surplus_token
                payment_amount = Decimal(0)  # All tokens used
                break

        # Update the teacher's token balance with the remaining amount
        cur.execute("""
            UPDATE teacher_token
            SET token_paid = %s
            WHERE teacher_id = %s AND term_id = %s
        """, (payment_amount + surplus_token, teacher_id, active_term_id))

        # Commit changes
        my_db.commit()

        # Inform the user
        messagebox.showinfo("Remedial App", "Payment processing completed successfully.")
        pay_token_entry.delete(0, END)
        display_teachers()

    except ValueError as ex:
        messagebox.showerror("Remedial App", f"Invalid input: {ex}\nPlease check your entries.")
    except Exception as ex:
        messagebox.showerror("Remedial App", f"An unexpected error occurred: {ex}")


def track_teacher_attendance():
    global teacher_tree,session_combo,grade_combo,subject_combo,pay_token_entry,tr_win_disp
    try:
        # Step 1: Get the active term ID
        cur.execute("""SELECT term_id FROM term WHERE is_active=1""")
        term = cur.fetchone()
        if not term:
            messagebox.showwarning("Remedial App", "No active term found. Please set an active term.")
            return
        term_id = term[0]

        # Step 2: Get the selected teacher ID from the treeview
        try:
            pos = teacher_tree.selection()
            value = teacher_tree.item(pos, "values")
            teacher_id = int(value[1])  # Extract and convert teacher_id
        except IndexError:
            messagebox.showerror("Remedial App", "Please select a teacher from the list.")
            return

        # Step 3: Get the active week number ID
        cur.execute("""SELECT week_number_id FROM week_number WHERE is_active=1 AND term_id=%s""", (term_id,))
        week_no = cur.fetchone()
        if not week_no:
            messagebox.showwarning("Remedial App", "No active week found. Please set an active week.")
            return
        week_number_id = week_no[0]

        # Step 4: Get the session and learning area
        session = session_combo.get()
        learning_area = subject_combo.get()
        if not session or not learning_area:
            messagebox.showerror("Remedial App", "Please select both the session and subject.")
            return

        # Step 5: Determine the session amount
        if session in ["Morning", "Evening"]:
            cur.execute("""SELECT weekday FROM termly_pay""")
        elif session == "Saturday":
            cur.execute("""SELECT weekend FROM termly_pay """)
        else:
            messagebox.showerror("Remedial App", "Invalid session type selected.")
            return

        pay = cur.fetchone()
        if not pay or pay[0] < 1:
            messagebox.showwarning("Remedial App", "Please set the weekly or weekend pay for the current term.")
            return
        session_amount = float(pay[0])

        # Step 6: Get the grade for the session
        tr_attend_grade = grade_combo.get()
        if not tr_attend_grade:
            messagebox.showerror("Remedial App", "Please select a grade.")
            return

        # # Step 7: Retrieve the teacher's total and deduct any paid amounts
        # cur.execute("""
        #     SELECT teacher_total FROM teacher_attendance
        #     WHERE teacher_id=%s ORDER BY teacher_attendance_id DESC LIMIT 1
        # """, (teacher_id,))
        # prev = cur.fetchone()
        # prev_total = float(prev[0]) if prev else 0  # Use 0 if no previous total

        # # Check if the teacher has any payment recorded in teacher_token
        # cur.execute("""
        #     SELECT token_paid FROM teacher_token
        #     WHERE teacher_id=%s AND term_id=%s
        # """, (teacher_id, term_id))
        # payment = cur.fetchone()
        # payment_amount = float(payment[0]) if payment else 0

        # # Deduct the payment amount from the previous total
        # if payment_amount > 0:
        #     new_teacher_total = prev_total + session_amount - payment_amount

        #     # Clear the payment record to avoid repeated deductions
        #     cur.execute("""
        #         UPDATE teacher_token SET token_paid=0
        #         WHERE teacher_id=%s AND term_id=%s
        #     """, (teacher_id, term_id))
        #     my_db.commit()
        # else:
        #     new_teacher_total = prev_total + session_amount

        # Step 8: Insert or update attendance record
        if teacher_id and week_number_id and session and tr_attend_grade and session_amount and learning_area:
            cur.execute("""
                INSERT  INTO teacher_attendance
                    (teacher_id, week_number_id, session, grade, session_amount,  subject, term_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (teacher_id, week_number_id, session, tr_attend_grade, session_amount, learning_area, term_id))
            my_db.commit()
            tr_win_disp.configure(text="Teacher attendance recorded successfully")
            tr_win_disp.after(4000, lambda: tr_win_disp.configure(text=""))
            display_teachers()
        else:
            messagebox.showerror("Remedial App", "Submission failed. Please fill all required fields.")
            return

    except Exception as e:
        messagebox.showerror("Remedial App", f"Error: {e} Occurred")

#setting week one as default in each new term
def week_one():
    cur.execute("UPDATE week_number SET is_active=0 WHERE is_active=1 ")
    cur.execute("""UPDATE week_number SET is_active=1 WHERE selected_week='One' AND 
                term_id=(SELECT term_id FROM term WHERE is_active=1)""")
    my_db.commit()
    cur.execute("SELECT selected_week FROM week_number WHERE is_active=1")
    week_number=cur.fetchone()[0]
    week_combo.set(week_number)
def set_default_week(e=None):
    selected_week=week_combo.get()
    #deactivating the rest of terms
    cur.execute("UPDATE week_number SET  is_active=0 WHERE is_active=1")
    my_db.commit()
    #updating selected term to active
    cur.execute("""UPDATE week_number SET is_active=1  WHERE selected_week=%s AND
                term_id=(SELECT term_id FROM term WHERE is_active=1)""",(selected_week,))
    my_db.commit()
    # print(selected_week)
def retrieve_week():
    # retrieving selected term and setting it to default week
    cur.execute("""SELECT selected_week FROM week_number WHERE is_active=1
              AND term_id=(SELECT term_id FROM term WHERE is_active=1)""")
    default = cur.fetchone()
    if default:
        week_no = default[0]
        week_combo.set(week_no)
    else:
        week_combo.set(weeks[0] if weeks else '')


def promote_learners_func(e=None):
    global move_learner_combo, undo_move_combo
    from_grade = move_learner_combo.get()
    to_grade = undo_move_combo.get()

    if from_grade == "Move From Grade" or to_grade == "Move To":
        messagebox.showinfo("Remedial App", "Select both From and To Grade to proceed.")


        return

    if from_grade == to_grade:
        messagebox.showinfo("Remedial App", "From and To Grades cannot be the same.")
        return

    # Define valid adjacent moves
    valid_moves = {
        "Seven": ["Eight"],
        "Eight": ["Seven", "Nine"],
        "Nine": ["Eight", "Archive"],
        "Archive": ["Nine"]
    }

    # Check if the move is valid
    if to_grade not in valid_moves.get(from_grade, []):
        messagebox.showinfo("Remedial App", f"Invalid move! You cannot move from {from_grade} to {to_grade}.")
        return

    # Determine where to check for the grade
    if from_grade == "Archive":
        cur.execute("SELECT COUNT(*) FROM archive WHERE grade = %s", (to_grade,))
        grade_count = cur.fetchone()[0]
    else:
        cur.execute("SELECT COUNT(*) FROM learner WHERE grade = %s", (from_grade,))
        grade_count = cur.fetchone()[0]

    if grade_count == 0:
        messagebox.showinfo("Remedial App", f"The grade '{from_grade}' is not available in the database.")
        return

    try:
        # Promote learners within the learner table
        if from_grade == "Seven" and to_grade == "Eight":
            cur.execute("UPDATE learner SET grade='Eight' WHERE grade='Seven'")
        elif from_grade == "Eight" and to_grade == "Seven":
            cur.execute("UPDATE learner SET grade='Seven' WHERE grade='Eight'")
        elif from_grade == "Eight" and to_grade == "Nine":
            cur.execute("UPDATE learner SET grade='Nine' WHERE grade='Eight'")
        elif from_grade == "Nine" and to_grade == "Eight":
            cur.execute("UPDATE learner SET grade='Eight' WHERE grade='Nine'")

        # Move Grade 9 learners to the archive
        elif from_grade == "Nine" and to_grade == "Archive":
            cur.execute("SELECT * FROM learner WHERE grade='Nine'")
            rows = cur.fetchall()

            if rows:
                for item in rows:
                    cur.execute("""
                        INSERT INTO archive(learner_id, first, second, surname, grade)
                        VALUES (%s, %s, %s, %s, %s)
                    """, (item[0], item[1], item[2], item[3], item[4]))

                cur.execute("DELETE FROM learner WHERE grade='Nine'")
                messagebox.showinfo("Move Learner", "Learners were successfully archived.")
            else:
                messagebox.showinfo("Remedial App", "No Grade Nine learners were found!")

        # Restore learners from archive to learner table
        elif from_grade == "Archive" and to_grade == "Nine":
            cur.execute("SELECT * FROM archive WHERE grade='Nine'")
            rows = cur.fetchall()

            if rows:
                for item in rows:
                    cur.execute("""
                        INSERT INTO learner(learner_id, first, second, surname, grade)
                        VALUES (%s, %s, %s, %s, %s)
                    """, (item[0], item[1], item[2], item[3], item[4]))

                cur.execute("DELETE FROM archive WHERE grade='Nine'")
                messagebox.showinfo("Move Learner", "Learners were successfully restored from archive.")
            else:
                messagebox.showinfo("Remedial App", "No Grade Nine learners were found in the archive!")

        my_db.commit()
        display_learners()

    except Exception as e:
        my_db.rollback()
        messagebox.showerror("Error", f"An error occurred: {e}")

#search function
def binding(e):
    # learner_menu.tk_popup(e.x,e.y)
    x, y = e.x_root, e.y_root  # Get the absolute mouse position
    x = max(0, min(x, e.widget.winfo_screenwidth() - 10))  # Ensure it stays within screen width
    y = max(0, min(y, e.widget.winfo_screenheight() - 10))  # Ensure it stays within screen height
    learner_menu.tk_popup(x, y)

    #deleting learners
def delete_transaction():
    try:
        adm = []  # List to hold learner_ids for deletion
        tuples = []  # List to hold parameters for database query
        pos = learner_tree.selection()  # Get selected rows in learner_tree

        if not pos:
            messagebox.showwarning("Remedial App", "Select records before attempting to delete")
        else:
            root.bell()
            resp = messagebox.askyesno("Remedial App", "Are sure you want to delete\nSelected PAYMENT record(s)\nPermanently?")
            
            if resp == 1:  # Proceed if user confirms
                for items in pos:
                    values = learner_tree.item(items, "values")
                    adm.append(values[1])  # Extract learner_id from the tree view
                    learner_tree.delete(items)  # Delete the item from the learner_tree

                # Prepare the tuples for the SQL query
                for a in adm:
                    tuples.append((a,))

                # Step 1: Fetch the most recent transaction_id for each learner
                for learner_id in adm:
                    cur.execute("""
                        SELECT t.transaction_id
                        FROM transactions t
                        JOIN transaction_history th ON t.transaction_id = th.transaction_id
                        WHERE t.learner_id = %s
                        AND t.term_id = (SELECT term_id FROM term WHERE is_active = 1)
                        ORDER BY th.trans_time DESC
                        LIMIT 1
                    """, (learner_id,))  # Pass each learner_id one by one

                    # Step 2: Fetch the transaction_id to delete
                    transaction_id_to_delete = cur.fetchone()

                    # Step 3: If transaction_id exists, delete it
                    if transaction_id_to_delete:
                        cur.execute("""
                            DELETE FROM transactions
                            WHERE transaction_id = %s
                        """, (transaction_id_to_delete[0],))

                        my_db.commit()  # Commit the changes
                        display_learners()
            else:
                pass  # If user cancels, do nothing
    except Exception as e:
        messagebox.showerror("Remedial App"f"{e}")

#deleting learners
def delete_learner():
        adm=[]
        tuples=[]
        pos=learner_tree.selection()
        if not pos:
            messagebox.showwarning("Remedial App","Select records before attempting to delete")
        else:
            root.bell()
            resp=messagebox.askyesno("Remedial App","Are sure you want to delete\n LEARNER(S) permanently?")
            if resp==1:
                for items in pos:
                    values=learner_tree.item(items,"values") 
                    adm.append(values[1])
                    #deleting from learner_tree
                    learner_tree.delete(items)
                for a in adm:
                    tuples.append((a,))
                cur.executemany("DELETE FROM learner WHERE learner_id=%s",tuples)
                my_db.commit()
            else:
                pass

# Treeview to display results

def search_func(e=None):
    search_term = learner_search_entry.get().strip().lower()
    search_by = search_by_combo.get()
    selected_grade = disp_bal_combo.get().strip().lower()

    if search_by == "Search By":
        messagebox.showinfo("Select Criteria", "Please select a search criteria before searching.")
        return

    if not search_term:
        messagebox.showwarning("Input Error", "Search entry cannot be blank.")
        return

    try:
        learner_tree.delete(*learner_tree.get_children())

        if search_by == "Adm No":
            try:
                search_term = int(search_term)
            except ValueError:
                messagebox.showwarning("Input Error", "Admission Number must be a number.")
                return

            if selected_grade == "display all learners":
                query = """
                SELECT 
                    l.learner_id,
                    l.grade,
                    l.first,
                    l.second,
                    l.surname,
                    COALESCE(SUM(t.amount_paid) OVER (PARTITION BY l.learner_id), 0) AS amount_paid,
                    COALESCE((SELECT lnr_pay FROM termly_pay LIMIT 1) - 
                             SUM(t.amount_paid) OVER (PARTITION BY l.learner_id), 0) AS balance
                FROM learner l
                LEFT JOIN transactions t 
                    ON l.learner_id = t.learner_id 
                    AND t.term_id = (SELECT term_id FROM term WHERE is_active = 1)
                WHERE l.learner_id = %s
                """
                cur.execute(query, (search_term,))
            else:
                query = """
                SELECT 
                    l.learner_id,
                    l.grade,
                    l.first,
                    l.second,
                    l.surname,
                    COALESCE(SUM(t.amount_paid) OVER (PARTITION BY l.learner_id), 0) AS amount_paid,
                    COALESCE((SELECT lnr_pay FROM termly_pay LIMIT 1) - 
                             SUM(t.amount_paid) OVER (PARTITION BY l.learner_id), 0) AS balance
                FROM learner l
                LEFT JOIN transactions t 
                    ON l.learner_id = t.learner_id 
                    AND t.term_id = (SELECT term_id FROM term WHERE is_active = 1)
                WHERE l.learner_id = %s AND LOWER(l.grade) = %s
                """
                cur.execute(query, (search_term, selected_grade))
        else:
            if selected_grade == "display all learners":
                query = """
                SELECT 
                    l.learner_id,
                    l.grade,
                    l.first,
                    l.second,
                    l.surname,
                    COALESCE(SUM(t.amount_paid) OVER (PARTITION BY l.learner_id), 0) AS amount_paid,
                    COALESCE((SELECT lnr_pay FROM termly_pay LIMIT 1) - 
                             SUM(t.amount_paid) OVER (PARTITION BY l.learner_id), 0) AS balance
                FROM learner l
                LEFT JOIN transactions t 
                    ON l.learner_id = t.learner_id 
                    AND t.term_id = (SELECT term_id FROM term WHERE is_active = 1)
                WHERE LOWER(l.first) LIKE %s OR LOWER(l.second) LIKE %s OR LOWER(l.surname) LIKE %s
                """
                search_pattern = f"%{search_term}%"
                cur.execute(query, (search_pattern, search_pattern, search_pattern))
            else:
                query = """
                SELECT 
                    l.learner_id,
                    l.grade,
                    l.first,
                    l.second,
                    l.surname,
                    COALESCE(SUM(t.amount_paid) OVER (PARTITION BY l.learner_id), 0) AS amount_paid,
                    COALESCE((SELECT lnr_pay FROM termly_pay LIMIT 1) - 
                             SUM(t.amount_paid) OVER (PARTITION BY l.learner_id), 0) AS balance
                FROM learner l
                LEFT JOIN transactions t 
                    ON l.learner_id = t.learner_id 
                    AND t.term_id = (SELECT term_id FROM term WHERE is_active = 1)
                WHERE (LOWER(l.first) LIKE %s OR LOWER(l.second) LIKE %s OR LOWER(l.surname) LIKE %s) 
                  AND LOWER(l.grade) = %s
                """
                search_pattern = f"%{search_term}%"
                cur.execute(query, (search_pattern, search_pattern, search_pattern, selected_grade))

        results = cur.fetchall()

        if results:
            for index, row in enumerate(results, start=1):
                learner_tree.insert(
                    "",
                    "end",
                    values=(
                        index,
                        row[0],  # Learner ID
                        row[1],  # Grade
                        f"{row[2].title()} {row[3].title()} {row[4].title()}",
                        row[5],  # Amount Paid
                        row[6]   # Balance
                    )
                )
        else:
            display_learners()
            search_label.configure(text="Not found")
            search_label.after(4000, lambda: search_label.configure(text=""))

    except Exception as e:
        messagebox.showerror("Database Error", f"An error occurred: {e}")
    finally:
        learner_search_entry.delete(0, END)


# def make_payment(e=None):
#     try:
#         # Select learner ID from the learner treeview
#         pos = learner_tree.selection()
#         adm = learner_tree.item(pos, "values")
#         learner_id = adm[1]

#         amount_paid = float(amount_entry.get())

#         # Retrieve the current term's number
#         cur.execute("SELECT term_number FROM term WHERE is_active = 1")
#         term_no = cur.fetchone()
#         if not term_no:
#             messagebox.showwarning("Remedial App", "No active term found. Please check term setup.")
#             return
#         current_term_number = term_no[0]

#         # Determine the previous term ID (if applicable)
#         prev_term_id = None if current_term_number == 1 else cur.execute(
#             "SELECT term_id FROM term WHERE term_number = %s", (current_term_number - 1,)
#         ).fetchone()
#         prev_term_id = prev_term_id[0] if prev_term_id else None

#         # Check if learner has outstanding balance from the previous term
#         if prev_term_id:
#             cur.execute("SELECT amount_paid, balance FROM transactions WHERE learner_id = %s", (learner_id,))
#             balance = cur.fetchone()
#             prev_balance = balance[1] if balance else 0
#             prev_amount_paid = balance[0] if balance else 0

#             if prev_balance > 0 or prev_amount_paid == 0:
#                 cur.execute("SELECT first, second, surname FROM learner WHERE learner_id = %s", (learner_id,))
#                 name = cur.fetchone()
#                 fullname = f"\n{name[0].title()} {name[1].title()} {name[2].title()}"
#                 messagebox.showwarning(
#                     "Remedial App",
#                     f"{fullname} has not cleared\ntheir previous term fees. Amount paid so\nfar is: KSH {prev_amount_paid:.2f}."
#                 )
#                 return

#         # Retrieve the termly fee (lnr_pay) for the learner's grade
#         cur.execute("SELECT lnr_pay FROM termly_pay LIMIT 1")
#         amt = cur.fetchone()
#         tot = float(amt[0]) if amt else 0

#         if tot < 1:
#             cur.execute("SELECT selected_term FROM term WHERE is_active = 1")
#             is_active = cur.fetchone()[0]
#             messagebox.showwarning("Remedial App", f"Set Fee payable for {is_active}")
#             return

#         # Check for existing transaction records for the current term
#         cur.execute("""SELECT amount_paid, balance FROM transactions
#                         WHERE learner_id = %s AND term_id = (SELECT term_id FROM term WHERE is_active = 1)""",
#                     (learner_id,))
#         record = cur.fetchone()

#         if record:
#             new_amount = float(record[0]) + amount_paid
#             new_balance = float(record[1]) - amount_paid
#         else:
#             new_amount = amount_paid
#             new_balance = tot - amount_paid

#         # Ensure the amount paid doesn't exceed the termly fee
#         if new_amount > tot:
#             messagebox.showwarning("Remedial App", "Amount Paid cannot exceed termly\npayable amount")
#             return

#         # Retrieve the current term ID
#         cur.execute("SELECT term_id FROM term WHERE is_active = 1")
#         active = cur.fetchone()[0]

#         # Insert or update the transaction
#         cur.execute("""INSERT INTO transactions(amount_paid, learner_id, term_id, balance)
#                         VALUES (%s, %s, %s, %s)
#                         ON DUPLICATE KEY UPDATE
#                         amount_paid = VALUES(amount_paid),
#                         balance = VALUES(balance)""", (new_amount, learner_id, active, new_balance))
#         my_db.commit()

#         # Record transaction history
#         comment = (From_entry.get()).title()
#         if len(comment) > 25:
#             raise ValueError("Comment exceeds 25 characters")
#         cur.execute("SELECT transaction_id FROM transactions ORDER BY transaction_id DESC LIMIT 1")
#         item = cur.fetchone()
#         transaction_id = item[0] if item else None

#         cur.execute("""INSERT INTO transaction_history(learner_id, amount, balance, term_id, comment, transaction_id)
#                     VALUES (%s, %s, %s, %s, %s, %s)""",
#                     (learner_id, new_amount, new_balance, active, comment, transaction_id))
#         my_db.commit()

#         # Clean up old records if the balance is cleared
#         cur.execute("""DELETE FROM transaction_history WHERE learner_id = (
#                         SELECT learner_id FROM transactions WHERE balance = 0 AND learner_id = %s 
#                         AND term_id = (SELECT term_id FROM term WHERE is_active = 1))""", (learner_id,))
#         my_db.commit()

#         # Clear input and notify success
#         amount_entry.delete(0, END)
#         disp_label.configure(text="Payment saved successfully")
#         disp_label.after(4000, lambda: disp_label.configure(text=""))
#         display_learners()
#     except Exception as ex:
#         messagebox.showerror("Remedial App", f"An error occurred: {ex}")

def make_payment(e=None):
    try:
        # Select learner ID from the learner treeview
        pos = learner_tree.selection()
        adm = learner_tree.item(pos, "values")
        learner_id = adm[1]

        amount_paid = float(amount_entry.get())

        # Retrieve the current term's number
        cur.execute("SELECT term_number FROM term WHERE is_active = 1")
        term_no = cur.fetchone()
        if not term_no:
            messagebox.showwarning("Remedial App", "No active term found. Please check term setup.")
            return
        current_term_number = term_no[0]

        # Determine the previous term ID (if applicable)
        prev_term_id = None
        if current_term_number > 1:
            cur.execute("SELECT term_id FROM term WHERE term_number = %s", (current_term_number - 1,))
            prev_term_id = cur.fetchone()
            prev_term_id = prev_term_id[0] if prev_term_id else None

        # Check if learner has outstanding balance from the previous term
        if prev_term_id:
            cur.execute("SELECT amount_paid, balance FROM transactions WHERE learner_id = %s AND term_id = %s", (learner_id, prev_term_id))
            balance = cur.fetchone()
            prev_balance = balance[1] if balance else 0
            prev_amount_paid = balance[0] if balance else 0

            if prev_balance > 0 or prev_amount_paid == 0:
                cur.execute("SELECT first, second, surname FROM learner WHERE learner_id = %s", (learner_id,))
                name = cur.fetchone()
                fullname = f"\n{name[0].title()} {name[1].title()} {name[2].title()}"
                messagebox.showwarning(
                    "Remedial App",
                    f"{fullname} has not cleared\ntheir previous term fees. Amount paid so\nfar is: KSH {prev_amount_paid:.2f}."
                )
                return

        # Retrieve the termly fee (lnr_pay) for the learner's grade
        cur.execute("SELECT lnr_pay FROM termly_pay LIMIT 1")
        amt = cur.fetchone()
        tot = float(amt[0]) if amt else 0

        if tot < 1:
            cur.execute("SELECT selected_term FROM term WHERE is_active = 1")
            is_active = cur.fetchone()[0]
            messagebox.showwarning("Remedial App", f"Set Fee payable for {is_active}")
            return

        # Check for existing transaction records for the current term
        cur.execute("""SELECT amount_paid, balance FROM transactions
                        WHERE learner_id = %s AND term_id = (SELECT term_id FROM term WHERE is_active = 1)""",
                    (learner_id,))
        record = cur.fetchone()

        if record:
            new_amount = float(record[0]) + amount_paid
            new_balance = float(record[1]) - amount_paid
        else:
            new_amount = amount_paid
            new_balance = tot - amount_paid

        # Ensure the amount paid doesn't exceed the termly fee
        if new_amount > tot:
            messagebox.showwarning("Remedial App", "Amount Paid cannot exceed termly\npayable amount")
            return

        # Retrieve the current term ID
        cur.execute("SELECT term_id FROM term WHERE is_active = 1")
        active = cur.fetchone()[0]

        # Insert or update the transaction
        cur.execute("""INSERT INTO transactions(amount_paid, learner_id, term_id, balance)
                        VALUES (%s, %s, %s, %s)
                        ON DUPLICATE KEY UPDATE
                        amount_paid = VALUES(amount_paid),
                        balance = VALUES(balance)""", (new_amount, learner_id, active, new_balance))
        my_db.commit()

        # Record transaction history
        comment = (From_entry.get()).title()
        if len(comment) > 25:
            raise ValueError("Comment exceeds 25 characters")
        cur.execute("SELECT transaction_id FROM transactions ORDER BY transaction_id DESC LIMIT 1")
        item = cur.fetchone()
        transaction_id = item[0] if item else None

        cur.execute("""INSERT INTO transaction_history(learner_id, amount, balance, term_id, comment, transaction_id)
                    VALUES (%s, %s, %s, %s, %s, %s)""",
                    (learner_id, new_amount, new_balance, active, comment, transaction_id))
        my_db.commit()

        # Clean up old records if the balance is cleared
        cur.execute("""DELETE FROM transaction_history WHERE learner_id = (
                        SELECT learner_id FROM transactions WHERE balance = 0 AND learner_id = %s 
                        AND term_id = (SELECT term_id FROM term WHERE is_active = 1))""", (learner_id,))
        my_db.commit()

        # Clear input and notify success
        amount_entry.delete(0, END)
        disp_label.configure(text="Payment saved successfully")
        disp_label.after(4000, lambda: disp_label.configure(text=""))
        display_learners()
    except Exception as ex:
        messagebox.showerror("Remedial App", f"An error occurred: {ex}")
def disable_combo(e):
    global tr_title_combo,reg_grade_combo,person_type_combo
    if person_type_combo.get()=="Learner":
        tr_title_combo.configure(state=DISABLED)
        reg_grade_combo.configure(state=NORMAL)
    else:
        if person_type_combo.get()=="Teacher":
            tr_title_combo.configure(state=NORMAL)
            reg_grade_combo.configure(state=DISABLED)


def display_teachers(e=None):
    global teacher_tree
    try:
        # Clear the tree before inserting new data
        teacher_tree.delete(*teacher_tree.get_children())

        # Retrieve teacher details along with the token paid amount and total amount owed
        cur.execute("""
            SELECT t.teacher_id, t.title, t.first, t.second, t.surname, 
                IFNULL((SELECT SUM(tt.token_paid) FROM teacher_token tt WHERE tt.teacher_id = t.teacher_id), 0) AS total_paid,
                IFNULL((SELECT SUM(ta.session_amount) FROM teacher_attendance ta WHERE ta.teacher_id = t.teacher_id), 0) AS total_owed
            FROM teacher t
            GROUP BY t.teacher_id
        """)
        teachers = cur.fetchall()

        if teachers:
            for index, teacher in enumerate(teachers, start=1):
                teacher_id, title, first, second, surname, total_paid, total_owed = teacher

                # Calculate balance (owed - paid)
                balance = total_owed - total_paid

                # Insert teacher details into the Treeview with numbering, token paid amount, and balance
                teacher_tree.insert(
                    "", "end",
                    values=(index, teacher_id, f"{title.title()} {first.title()} {second.title()} {surname.title()}",
                            total_paid, total_owed, balance)
                )
        else:
            messagebox.showinfo("Remedial App", "No records for teachers")
    except Exception as e:
        messagebox.showerror("Remedial App", f"Error: {e}")

def display_learners(e=None):
    try:
        selected_option = disp_bal_combo.get()

        # Get the current term's lnr_pay without linking to term_id
        cur.execute("SELECT lnr_pay FROM termly_pay LIMIT 1")
        termly_pay = cur.fetchone()
        lnr_pay = termly_pay[0] if termly_pay else 0  # Default to 0 if no value exists

        # Determine if displaying all learners or learners from a specific grade
        if selected_option == "Display all Learners":
            cur.execute("""
            SELECT 
                learner.learner_id, 
                learner.grade, 
                learner.first, 
                learner.second, 
                learner.surname, 
                IFNULL(SUM(transactions.amount_paid), 0) AS amount_paid, 
                %s - IFNULL(SUM(transactions.amount_paid), 0) AS balance
            FROM learner
            LEFT JOIN transactions 
                ON learner.learner_id = transactions.learner_id 
                AND transactions.term_id = (SELECT term_id FROM term WHERE is_active = 1)
            GROUP BY learner.learner_id
            ORDER BY amount_paid DESC
            """, (lnr_pay,))
        else:
            cur.execute("""
            SELECT 
                l.learner_id,
                l.grade,
                l.first,
                l.second,
                l.surname,
                IFNULL(SUM(t.amount_paid), 0) AS amount_paid,
                %s - IFNULL(SUM(t.amount_paid), 0) AS balance
            FROM learner l
            LEFT JOIN transactions t 
                ON l.learner_id = t.learner_id 
                AND t.term_id = (SELECT term_id FROM term WHERE is_active = 1)
            WHERE l.grade = %s
            GROUP BY l.learner_id
            ORDER BY amount_paid DESC
            """, (lnr_pay, selected_option))

        learners = cur.fetchall()

        # Clear the treeview before inserting new data
        learner_tree.delete(*learner_tree.get_children())

        if learners:
            for index, learner in enumerate(learners, start=1):
                learner_tree.insert(
                    "",
                    "end",
                    values=(
                        index,
                        learner[0],  # Learner ID
                        learner[1],  # Grade
                        f"{learner[2].title()} {learner[3].title()} {learner[4].title()}",  # Full Name
                        learner[5],  # Amount Paid
                        learner[6]   # Balance
                    )
                )
        else:
            messagebox.showinfo("Remedial App", f"No records found for {'any learners' if selected_option == 'Display all Learners' else 'grade ' + selected_option}.")

    except Exception as e:
        messagebox.showerror("Remedial App", f"An error occurred: {e}")

                 
#clear person_win boxes
def clear_person_win():
    global number_entry,first_entry,second_entry,surname_entry
    number_entry.delete(0,END)
    first_entry.delete(0,END)
    second_entry.delete(0,END)
    surname_entry.delete(0,END)
#clearing fields in term win
def clear_term_win():
    global term_combo,set_term_label,lnr_amount_entry,tr_weekend_entry,tr_week_entry
    lnr_amount_entry.delete(0,END)
    tr_weekend_entry.delete(0,END)
    tr_week_entry.delete(0,END)

def clear_learner_tree():
    learner_tree.delete(*learner_tree.get_children())
#deleting function(learner,teacher)
def delete_person(person):  
    try:
        global person_type_combo,number_entry,reg_grade_combo,tr_title_combo,first_entry,second_entry,surname_entry,reg_person_label 
        id_number=number_entry.get()
        id_number=int(id_number)
        person=person_type_combo.get()
        if not person:
            root.bell()
            # messagebox.showwarning("Register Persons","Select person type to proceed")
            # CTkMessagebox(title="Register Persons",message="Select person type to proceed",icon="warning")
            messagebox.showwarning(title="Register Persons",message="Select person type to proceed")
            return
        #deleting learner
        if person=="Learner":
            #access learner before asking to delete
            cur.execute("""SELECT first,second,surname FROM learner WHERE learner_id=%s
                        """,(id_number,))
            lnr=cur.fetchone()
            if lnr:
                learner=f"{lnr[0]} {lnr[1]} {lnr[2]} "
            else:
                reg_person_label.configure(text=f"Learner ID {id_number} not found")
                reg_person_label.after(4000,lambda:reg_person_label.configure(text=""))
                return
            if learner:
                root.bell()
                resp=messagebox.askyesno("Register Persons",f"Are you sure you want to delete\n{learner}?")
                if resp:
                    cur.execute("DELETE FROM learner WHERE learner_id=%s",(id_number,))
                    my_db.commit()
                    clear_learner_tree()
                    display_learners()
                    reg_person_label.configure(text="Learner details deleted successfully")
                    reg_person_label.after(4000,lambda:reg_person_label.configure(text=""))
                    number_entry.delete(0,END)
        else:
            person=="Teacher"
            #access learner before asking to delete
            cur.execute("""SELECT title,first,second,surname FROM teacher WHERE teacher_id=%s
                        """,(id_number,))
            tr=cur.fetchone()
            if tr:
                teacher=f"{tr[0]} {tr[1]} {tr[2]} {tr[3]} "
            else:
                reg_person_label.configure(text=f"Teacher ID {id_number} not found")
                reg_person_label.after(4000,lambda:reg_person_label.configure(text=""))
                return
            if teacher:
                root.bell()
                resp=messagebox.askyesno("Register Persons",f"Are you sure you want to delete\n{teacher}?")
                if resp:
                    cur.execute("DELETE FROM teacher WHERE teacher_id=%s",(id_number,))
                    my_db.commit()
                    reg_person_label.configure(text="Teacher details deleted successfully")
                    reg_person_label.after(4000,lambda:reg_person_label.configure(text=""))
                    number_entry.delete(0,END)
    except Exception as ex:
        messagebox.showerror("Register Persons",f"{ex}")
    
# 
def add_person(person):
    try:
        global person_type_combo, number_entry, reg_grade_combo, tr_title_combo, first_entry, second_entry, surname_entry, reg_person_label
        person = person_type_combo.get()
        id_number = number_entry.get()

        # Ensure ID is an integer
        try:
            id_number = int(id_number)
        except ValueError:
            messagebox.showwarning(title="Register Persons", message="ID number must be an integer.")
            return

        grade = reg_grade_combo.get()
        title = tr_title_combo.get()
        first = first_entry.get().strip().title()
        second = second_entry.get().strip().title()
        surname = surname_entry.get().strip().title()

        if not person:
            root.bell()
            CTkMessagebox(title="Register Persons", message="Select Person type to proceed", icon="warning")
            return

        if person == "Learner":
            if all([id_number, grade, first, second]):
                # Checking whether learner ID exists
                cur.execute("SELECT learner_id FROM learner WHERE learner_id=%s", (id_number,))
                learner_id = cur.fetchone()
                learner_id = learner_id[0] if learner_id else None

                # Inserting learner if ID does not exist
                if not learner_id:
                    cur.execute(
                        "INSERT INTO learner (learner_id, first, second, surname, grade) VALUES (%s, %s, %s, %s, %s)",
                        (id_number, first, second, surname, grade),
                    )
                    my_db.commit()
                    reg_person_label.configure(text="Learner details saved successfully")
                    reg_person_label.after(4000, lambda: reg_person_label.configure(text=""))
                    clear_person_win()
                    clear_learner_tree()
                    display_learners()
                # Updating learner details if ID exists
                else:
                    root.bell()
                    resp = messagebox.askyesno(
                        "Register Persons", f"Learner ID {id_number} exists. Do you want to overwrite?"
                    )
                    if resp:
                        cur.execute(
                            "UPDATE learner SET first=%s, second=%s, surname=%s, grade=%s WHERE learner_id=%s",
                            (first, second, surname, grade, id_number),
                        )
                        my_db.commit()
                        reg_person_label.configure(text="Learner details updated successfully")
                        clear_learner_tree()
                        display_learners()
                        clear_person_win()
                        reg_person_label.after(4000, lambda: reg_person_label.configure(text=""))
            else:
                root.bell()
                messagebox.showwarning(
                    "Register Persons", "Ensure that you have filled all the necessary fields for person type: Learner"
                )

        elif person == "Teacher":
            if all([id_number, title, first, second]):
                # Checking whether teacher ID exists
                cur.execute("SELECT teacher_id FROM teacher WHERE teacher_id=%s", (id_number,))
                teacher_id = cur.fetchone()
                teacher_id = teacher_id[0] if teacher_id else None

                # Inserting teacher record if ID does not exist
                if not teacher_id:
                    cur.execute(
                        "INSERT INTO teacher (teacher_id, title, first, second, surname) VALUES (%s, %s, %s, %s, %s)",
                        (id_number, title, first, second, surname),
                    )
                    my_db.commit()
                    reg_person_label.configure(text="Teacher details saved successfully")
                    reg_person_label.after(4000, lambda: reg_person_label.configure(text=""))
                    clear_person_win()
                else:
                    root.bell()
                    resp = messagebox.askyesno(
                        "Register Persons", f"Teacher ID {id_number} exists. Do you want to overwrite?"
                    )
                    if resp:
                        cur.execute(
                            "UPDATE teacher SET title=%s, first=%s, second=%s, surname=%s WHERE teacher_id=%s",
                            (title, first, second, surname, id_number),
                        )
                        my_db.commit()
                        reg_person_label.configure(text="Teacher details updated successfully")
                        clear_person_win()
                        reg_person_label.after(4000, lambda: reg_person_label.configure(text=""))
            else:
                root.bell()
                messagebox.showwarning(
                    "Register Persons", "Ensure that you have filled all the necessary fields for person type: Teacher"
                )
        else:
            messagebox.showwarning(title="Register Persons", message="Invalid person type selected.")
    except Exception as ex:
        messagebox.showwarning(title="Register Persons", message=f"An error occurred: {ex}")
        
#accessing grades
def grades_list():
    grades=["Display all Learners","Seven","Eight","Nine"]
    return grades
    
#terms function
def term_list():
    terms=["Term One,2025",
        "Term Two,2025",
        "Term Three, 2025",
        "Term One,2026",
        "Term Two,2026",
        "Term Three,2026",
        "Term One,2027",
        "Term Two,2027",
        "Term Three,2027",
        "Term One,2028",
        "Term Two,2028",
        "Term Three,2028",
        "Term One,2029",
        "Term Two,2029",
        "Term Three,2029",
        "Term One,2030",
        "Term Two,2030",
        "Term Three,2030"]
    return terms


def disp_pay_hist(e=None):  
    global search_hist_entry,pay_hist_tree,search_hist_disp
    pay_hist_tree.delete(*pay_hist_tree.get_children()) 
    try:       
        learner_id=search_hist_entry.get()
        learner_id=int(learner_id)
        
    #displaying transaction history
        #displaying all history
        cur.execute("""SELECT l.learner_id,l.grade,l.first,l.second,l.surname,
                    th.amount,th.balance,th.trans_time,th.comment FROM learner l
                    JOIN transaction_history th ON l.learner_id=th.learner_id
                     WHERE l.learner_id=%s ORDER BY trans_time DESC""",(learner_id,))
        h=cur.fetchall()
        if h:
            for index,h in enumerate(h,start=1):
                time=h[7]
                formatted_date=time.strftime("%d-%m-%Y")
                pay_hist_tree.insert("",END,values=(index,h[0],h[1],f"{h[2]} {h[3]} {h[4]}".title(),h[5],h[6],formatted_date,h[8]))
        else:
            search_hist_disp.configure(text=f"Record with Adm {search_hist_entry.get()} not found")
            search_hist_disp.after(4000,lambda:search_hist_disp.configure(text=""))
            call_pay_hist()
            search_hist_entry.delete(0,END)
    except Exception as e:
        messagebox.showerror("Remedial App",f"{e}")
        call_pay_hist()
# #setting termly pay
# #windows functions

def call_pay_hist():
    global pay_hist_tree
    cur.execute("""SELECT l.learner_id, l.grade, l.first, l.second, l.surname,
                   th.amount, th.balance, th.trans_time, th.comment FROM learner l
                   JOIN transaction_history th ON l.learner_id = th.learner_id
                   ORDER BY trans_time DESC""")
    h = cur.fetchall()
    # Insert fetched records into the treeview
    for index, record in enumerate(h, start=1):
        time = record[7]
        formatted_date = time.strftime("%d-%m-%Y")
        pay_hist_tree.insert("", END, values=(index, record[0], record[1], f"{record[2]} {record[3]} {record[4]}",
                                              record[5], record[6], formatted_date, record[8]))
    # Check if records exist, else return
def payment_history_func(e=None):
    global pay_hist_tree, search_hist_entry, search_hist_disp

    # Fetch transaction history from the database
    cur.execute("""SELECT l.learner_id, l.grade, l.first, l.second, l.surname,
                th.amount, th.balance, th.trans_time, th.comment FROM learner l
                JOIN transaction_history th ON l.learner_id = th.learner_id
                ORDER BY trans_time DESC""")
    h = cur.fetchall()

    # Check if records exist, else return
    if not h:
        messagebox.showinfo("Remedial App", "No transaction history records")
        return

    # Create the window only if there are records
    pay_hist_win = ctk.CTkToplevel(root)
    pay_hist_win.geometry("810x400+400+10")
    pay_hist_win.resizable(width=False, height=False)
    pay_hist_win.transient(root)
    pay_hist_win.grab_set()
    pay_hist_win.title("Learner payment History")
    pay_hist_win.after(100, lambda: pay_hist_win.lift())  # Delay lifting the window
    pay_hist_win.after(200, lambda: pay_hist_win.focus_force())  # Delay forcing focus

    pay_hist_tree_frame = ctk.CTkFrame(pay_hist_win)
    pay_hist_tree_scroll = Scrollbar(pay_hist_tree_frame, orient=VERTICAL)
    pay_hist_tree_scroll.pack(side=RIGHT, fill=Y)
    pay_hist_tree_frame.place(x=20, y=50)

    pay_hist_tree = ttk.Treeview(pay_hist_tree_frame, yscrollcommand=pay_hist_tree_scroll.set, height=12,
                                 selectmode="extended")
    pay_hist_tree_scroll.configure(command=pay_hist_tree.yview)
    pay_hist_tree["columns"] = ("no", "ID", "grade", "name", "paid", "balance", "date", "comment")
    pay_hist_tree.column("#0", width=0, stretch=NO)
    pay_hist_tree.column("no", width=40, anchor="center", minwidth=35)
    pay_hist_tree.column("ID", width=40, minwidth=80, anchor=CENTER)
    pay_hist_tree.column("grade", width=80, minwidth=60, anchor=CENTER)
    pay_hist_tree.column("name", width=180, minwidth=100, anchor=W)
    pay_hist_tree.column("paid", width=80, minwidth=75, anchor=CENTER)
    pay_hist_tree.column("balance", width=60, minwidth=75, anchor=CENTER)
    pay_hist_tree.column("date", width=100, minwidth=45, anchor=CENTER)
    pay_hist_tree.column("comment", width=180, minwidth=150, anchor=W)

    # Headings
    pay_hist_tree.heading("#0", text="")
    pay_hist_tree.heading("no", text="#", anchor=CENTER)
    pay_hist_tree.heading("ID", text="Adm", anchor=CENTER)
    pay_hist_tree.heading("grade", text="Grade", anchor=CENTER)
    pay_hist_tree.heading("name", text="Name", anchor=W)
    pay_hist_tree.heading("paid", text="Paid", anchor=CENTER)
    pay_hist_tree.heading("balance", text="Bal", anchor=CENTER)
    pay_hist_tree.heading("date", text="Date", anchor=CENTER)
    pay_hist_tree.heading("comment", text="Comment", anchor=CENTER)
    pay_hist_tree.pack()

    # Insert fetched records into the treeview
    for index, record in enumerate(h, start=1):
        time = record[7]
        formatted_date = time.strftime("%d-%m-%Y")
        pay_hist_tree.insert("", END, values=(index, record[0], record[1], f"{record[2]} {record[3]} {record[4]}",
                                              record[5], record[6], formatted_date, record[8]))

    pay_hist_w_frame = ctk.CTkFrame(pay_hist_win,fg_color="transparent")
    pay_hist_w_frame.place(x=20, y=10)
    search_hist_entry = ctk.CTkEntry(pay_hist_w_frame, font=("Helvetica", 16), width=100,
                                     border_color=blue, placeholder_text="Enter Adm")
    search_hist_entry.grid(row=0, column=0, sticky=W)
    search_hist_button = ctk.CTkButton(pay_hist_w_frame, text="Search", command=disp_pay_hist, width=60)
    search_hist_button.grid(row=0, column=1, padx=5)
    search_hist_disp = ctk.CTkLabel(pay_hist_w_frame, text="", font=("Helvetica", 16))
    search_hist_disp.grid(row=0, column=2)


#teacher display
def teacher_win_func(e=None):
    global teacher_tree,session_combo,grade_combo,subject_combo,pay_token_entry,tr_win_disp
    teacher_win=ctk.CTkToplevel(root)
    teacher_win.geometry("600x500+400+10")
    teacher_win.resizable(width=False,height=False)
    teacher_win.transient(root)
    teacher_win.grab_set()
    teacher_win.title("Monitor Tr. Attendance")
    teacher_win.after(100, lambda: teacher_win.lift())  # Delay lifting the window
    teacher_win.after(200, lambda: teacher_win.focus_force())  # Delay forcing 
    teacher_tree_frame=ctk.CTkFrame(teacher_win)
    teacher_tree_scroll=Scrollbar(teacher_tree_frame,orient=VERTICAL)
    teacher_tree_scroll.pack(side=RIGHT,fill=Y)
    teacher_tree_frame.place(x=20,y=30)
    teacher_tree=ttk.Treeview(teacher_tree_frame,yscrollcommand=learner_tree_scroll.set,height=12,
    selectmode="extended")
    teacher_tree_scroll.configure(command=teacher_tree.yview)
    teacher_tree["columns"]=("no","ID","name","paid","balance")
    teacher_tree.column("#0",width=0,stretch=NO)
    teacher_tree.column("no",width=40,anchor="center",minwidth=35)
    teacher_tree.column("ID",width=100,minwidth=80,anchor=CENTER)
    teacher_tree.column("name",width=250,minwidth=180,anchor=W)
    teacher_tree.column("paid",width=80,minwidth=75,anchor=CENTER)
    teacher_tree.column("balance",width=80,minwidth=75,anchor=CENTER)
    #headings
    teacher_tree.heading("#0",text="")
    teacher_tree.heading("no",text="#",anchor=CENTER)
    teacher_tree.heading("ID",text="ID",anchor=CENTER)
    teacher_tree.heading("name",text="NAME",anchor=CENTER)
    teacher_tree.heading("paid",text="PAID",anchor=CENTER)
    teacher_tree.heading("balance",text="BAL",anchor=CENTER)
    teacher_tree.pack()
    #tr_win_widgets session_combo,grade_combo,subject_combo,pay_token_entry
    tr_win_w_frame=ctk.CTkFrame(teacher_win,fg_color="transparent")
    tr_win_w_frame.place(x=20,y=320)
    session_label=ctk.CTkLabel(tr_win_w_frame,text="Session",font=("Helvetica",16))
    session_label.grid(row=0,column=0)
    session_combo=ctk.CTkComboBox(tr_win_w_frame,font=("Helvetica",16),width=110,state="readonly",button_color=blue,values=("Morning","Evening","Saturday"))
    session_combo.grid(row=0,column=1,padx=10)
    grade_label=ctk.CTkLabel(tr_win_w_frame,text="Grade",font=("Helvetica",16))
    grade_label.grid(row=1,column=0,pady=10)
    #accessing grades 
    grades=grades_list()
    grade_combo=ctk.CTkComboBox(tr_win_w_frame,font=("Helvetica",16),width=110,state="readonly",button_color=blue,values=("Seven","Eight","Nine"))
    grade_combo.grid(row=1,column=1,padx=10)
    subjects=["MATHS","ENG","KISW","INT","SST","AGRI","CAS","CRE","PTC","PPI"]
    subject_label=ctk.CTkLabel(tr_win_w_frame,text="Subject",font=("Helvetica",16))
    subject_label.grid(row=2,column=0)
    subject_combo=ctk.CTkComboBox(tr_win_w_frame,font=("Helvetica",16),width=110,state="readonly",button_color=blue,values=(subjects))
    subject_combo.grid(row=2,column=1,padx=10)
    track_button=ctk.CTkButton(tr_win_w_frame,text="Save Changes",width=50,command=track_teacher_attendance)
    track_button.grid(row=3,column=1,pady=10)
    pay_token_label=ctk.CTkLabel(tr_win_w_frame,text="PayToken",font=("helvetica",16))
    pay_token_label.grid(row=0,column=2,padx=(70,10))
    pay_token_entry=ctk.CTkEntry(tr_win_w_frame,font=("Helvetica",16),width=110
                                  ,border_color=blue)
    pay_token_entry.grid(row=0,column=3)
    pay_button = ctk.CTkButton(tr_win_w_frame, text="Save Changes", width=50,
                               command=pay_token)
    pay_button.grid(row=1, column=3, pady=10)
    tr_win_disp=ctk.CTkLabel(teacher_win,text="",font=("helvetica",16))
    tr_win_disp.place(x=250,y=450)
    #menu
    attend_hist_menu=tb.Menu(teacher_win)
    teacher_win.configure(menu=attend_hist_menu)
    file_menu=Menu(attend_hist_menu)
    attend_hist_menu.add_cascade(menu=file_menu,label="Records")
    file_menu.add_command(label="History",command=tr_attendance_func)
    file_menu.add_command(label="Archives", command=lambda:disp_attend_archive_func(None))
    display_teachers()
#setting term(term/year,termly pay teacher pay,promoting learnes)
def set_term_func():
    global term_combo,set_term_label,lnr_amount_entry,lnr_amount_entry,tr_weekend_entry,tr_week_entry,move_learner_combo,undo_move_combo

    #accessing terms from term func
    terms=term_list()
    set_term_win=ctk.CTkToplevel(root)
    path="D:/Tonniegifted/Remedial App/Resources/remedial convert.ico"
    set_term_win.iconbitmap(path)
    set_term_win.geometry("320x350+400+10")
    set_term_win.resizable(width=False,height=False)
    set_term_win.transient(root)
    set_term_win.grab_set()
    set_term_win.title("Set Term")
    set_term_win.after(100, lambda: set_term_win.lift())  # Delay lifting the window
    set_term_win.after(200, lambda: set_term_win.focus_force())  # Delay forcing focus
    #set term win widgets
    term_label=ctk.CTkLabel(set_term_win,text="Term",font=("Helvetica",16))
    term_combo=ctk.CTkComboBox(set_term_win,values=terms,width=180,height=30,
                               state="readonly",font=("Helvetica",16),
                               button_color=blue,border_color="blue",command=switch_term)
    term_label.place(x=20,y=20)
    term_combo.place(x=110,y=20)
    termly_pay_label=ctk.CTkLabel(set_term_win,text="Termly Pay",font=("Helvetica",16),
                                  )
    termly_pay_label.place(x=20, y=60)
    lnr_amount_entry=ctk.CTkEntry(set_term_win,width=180,height=30,
                               font=("Helvetica",16),border_color=blue
                               ,placeholder_text="Learner Termly pay")
    # grade_entry_combo.set("Select Grade")
    lnr_amount_entry.place(x=110,y=60)
    tr_week_label=ctk.CTkLabel(set_term_win,text="Weekday",font=("Helvetica",16))
    tr_week_label.place(x=20,y=100)
    tr_week_entry=ctk.CTkEntry(set_term_win,font=("Helvetica",16),width=180
                                  ,border_color=blue,placeholder_text="Teacher Weekly pay")
    tr_week_entry.place(x=110,y=100)
    tr_weekend_label=ctk.CTkLabel(set_term_win,text="Weekend",font=("Helvetica",16))
    tr_weekend_label.place(x=20,y=140)
    tr_weekend_entry=ctk.CTkEntry(set_term_win,font=("Helvetica",16),width=180
                                  ,border_color=blue,placeholder_text="Teacher Weekend pay"
                                  )
    tr_weekend_entry.place(x=110,y=140)

    
    button=ctk.CTkButton(set_term_win,text="Save Changes",command=lambda:set_default_term(None)
                         ,width=60)
    button.place(x=20,y=230)
    move_learners=ctk.CTkButton(set_term_win,text="Move Learners",command=lambda:promote_learners_func(None)
                         ,width=60)
    move_learners.place(x=150,y=230)
    move_learner_combo = ctk.CTkComboBox(set_term_win, height=30, width=140, font=("Helvetica", 12),
                                         border_color=blue, button_color=blue,
                                         values=("Seven", "Eight", "Nine","Archive"), state="readonly")
    move_learner_combo.set("Move From Grade")
    move_learner_combo.place(x=20, y=180)
    undo_move_combo = ctk.CTkComboBox(set_term_win, height=30, width=120, font=("Helvetica", 12),
                                      border_color=blue, button_color=blue,
                                      values=("Seven", "Eight", "Nine","Archive"), state="readonly")
    undo_move_combo.place(x=170, y=180)
    undo_move_combo.set("Move To")
    #
    
    set_term_label=ctk.CTkLabel(set_term_win,text="",
                                font=("helvetica",14))
    set_term_label.place(x=50,y=260)
    #setting the default term
    cur.execute("SELECT selected_term FROM term WHERE is_active=1")
    term=cur.fetchone()
    if term:
        selected_term=term[0]
        term_combo.set(selected_term)
    else:
        term_combo.set(term_combo.cget("values")[0])

#adding and deleting teachers and learners win func
def add_person_func():
    global person_type_combo,number_entry,reg_grade_combo,tr_title_combo,first_entry,second_entry,surname_entry,reg_person_label
    person_win=ctk.CTkToplevel(root)
    person_win.geometry("350x430+400+10")
    person_win.title("Learners Registration")
    person_win.resizable(width=False,height=False)
    person_win.transient(root)
    person_win.grab_set()
    person_win.iconbitmap(path)
    person_win.after(100, lambda: person_win.lift())  # Delay lifting the window
    person_win.after(200, lambda: person_win.focus_force())  # Delay forcing focus
    #add person widget
    person_type_label=ctk.CTkLabel(person_win,text="Person Type",font=("Helvetica",16))
    person_type_label.place(x=57,y=10)
    person_type_combo=ctk.CTkComboBox(person_win,height=30,width=155,
                                      border_color=blue,button_color=blue,font=("Helvetica",16),
                                      values=("Learner","Teacher"),state="readonly",command=disable_combo)
    person_type_combo.place(x=150,y=10)
    number_label=ctk.CTkLabel(person_win,text="Id No",font=("Helvetica",16))
    number_label.place(x=20,y=50)
    number_entry=ctk.CTkEntry(person_win,height=30,border_color=blue,width=180,
                                font=("helvetica",16))
    number_entry.place(x=125,y=50)
    reg_grade_label=ctk.CTkLabel(person_win,text="Grade",font=("helvetica",16),)
    reg_grade_label.place(x=20,y=90)
    #accessing values from the function
    grade=grades_list()
    reg_grade_combo=ctk.CTkComboBox(person_win,height=30,border_color=blue,width=180,
                                font=("helvetica",16),state="readonly",values=("Seven","Eight","Nine"),
                                button_color=blue)
    reg_grade_combo.place(x=125,y=90)
    tr_title_label=ctk.CTkLabel(person_win,text="Title",font=("helvetica",16))
    tr_title_label.place(x=20,y=130)
    tr_title_combo=ctk.CTkComboBox(person_win,height=30,width=180,font=("Helvetica",16),
                                      border_color=blue,button_color=blue,
                                      values=("Tr","Mr","Mrs","Miss"),state="readonly")
    tr_title_combo.place(x=125,y=130)
    first_label=ctk.CTkLabel(person_win,text="First Name",font=("Helvetica",16))
    first_label.place(x=20,y=170)
    first_entry=ctk.CTkEntry(person_win,height=30,border_color=blue,width=180,
                                font=("helvetica",16))
    first_entry.place(x=125,y=170)
    second_label=ctk.CTkLabel(person_win,text="Second Name",font=("Helvetica",16))
    second_label.place(x=20,y=210)
    second_entry=ctk.CTkEntry(person_win,height=30,border_color=blue,width=180,
                                font=("helvetica",16))
    second_entry.place(x=125,y=210)
    
    surname_label=ctk.CTkLabel(person_win,text="Surname",font=("Helvetica",16))
    surname_label.place(x=20,y=250)
    surname_entry=ctk.CTkEntry(person_win,height=30,border_color=blue,width=180,
                                font=("helvetica",16))
    surname_entry.place(x=125,y=250)
     
    add_person_button=ctk.CTkButton(person_win,text="Add"
                         ,width=60,command=lambda:add_person(None))
    add_person_button.place(x=80,y=300)
    
    delete_person_button=ctk.CTkButton(person_win,text="Delete"
                         ,width=60,command=lambda:delete_person(None))
    delete_person_button.place(x=220,y=300)

    reg_person_label=ctk.CTkLabel(person_win,text="",
                                font=("helvetica",14))
    reg_person_label.place(x=60,y=340)


def switch_term(e=None):
    global term_combo
    selected_term = term_combo.get()


    # Deactivate all terms
    cur.execute("UPDATE term SET is_active = 0 WHERE is_active = 1")
    my_db.commit()

    # Activate the selected term
    cur.execute("UPDATE term SET is_active = 1 WHERE selected_term = %s", (selected_term,))
    my_db.commit()

    # Verify if the update worked
    cur.execute("SELECT selected_term, is_active FROM term WHERE is_active = 1")
    result = cur.fetchall()


    set_term_label.configure(text=f"{selected_term}: Active")
    set_term_label.after(4000, lambda: set_term_label.configure(text=""))

    # Checking if termly_pay exists
    cur.execute("""SELECT COUNT(*) FROM termly_pay""")
    termly_pay = cur.fetchone()[0]

    if termly_pay == 0:
        messagebox.showwarning("Set Term", "No termly pay found for the current term")

    # Refresh learners
    display_learners()
    week_one()

#retrieving the set week


def set_default_term(e):
    try:
        global term_combo, set_term_label, lnr_amount_entry, tr_weekend_entry, tr_week_entry
        selected = term_combo.get()  # Get the selected term from the combo box

        # Get pay values
        lnr_pay = float(lnr_amount_entry.get())  # Teacher pay
        tr_weekly_pay = float(tr_week_entry.get())  # Weekly pay
        tr_weekend_pay = float(tr_weekend_entry.get())  # Weekend pay

        if all([lnr_pay, tr_weekly_pay, tr_weekend_pay, selected]):

            # Ensure only one record exists by deleting the previous one
            cur.execute("DELETE FROM termly_pay")

            # Insert the new record
            cur.execute("""
                INSERT INTO termly_pay (weekday, weekend, lnr_pay)
                VALUES (%s, %s, %s)
            """, (tr_weekly_pay, tr_weekend_pay, lnr_pay))

            my_db.commit()

            # Show success message and update UI
            set_term_label.configure(text="Pay was set successfully")
            set_term_label.after(4000, lambda: set_term_label.configure(text=""))
            clear_term_win()
            display_learners()

        else:
            # If no active term is found
            root.bell()
            messagebox.showwarning(title="Set Term", message="No active term found.")
            # week_one()
            # else:
        #     root.bell()
        #     CTkMessagebox(title="Set Term", message="Fill all fields to set term", icon="warning")
    except Exception as ex:
        # Catch any exceptions
        messagebox.showwarning(title="Set Term", message=f"An error occurred: {ex}")
        my_db.rollback()


#placer
# def placer(e):
#     cord=f"{e.x} x {e.y}"
#     disp_label.configure(text=cord)
#     disp_label.after(3000,lambda:disp_label.configure(text=""))
    
#switching between light and dark modes
def toggle():
    #to work on how app can remember mode and to change path
        if toggle_switch.get()==1:
    # if toggle_switch.get()==1:
            ctk.set_appearance_mode("dark")
            ctk.set_default_color_theme("blue")
            # style 
            style=tb.Style()
            style.configure('Treeview',
                        rowheight=20,
                        font="times 13",
                        foreground="#FFFFFF",
                        background="#2A2A2A")
            style.configure("Treeview.Heading",foreground="#FFFFFF",font="Times 13",background="#2A2A2A")
            style.map('Treeview',background=[("selected","#1E90FF")])
            learner_tree_frame.configure(bg_color="#2A2A2A")
        else:
            ctk.set_appearance_mode("light")
            ctk.set_default_color_theme("blue")
            # style 
            style=tb.Style()
            style.configure('Treeview',
                        rowheight=20,
                        font="times 13",
                        background="#FFFFFF",
                        foreground="black")
            style.configure("Treeview.Heading",foreground="blue",font="Times 13",background="#FFFFFF")
            style.map('Treeview',background=[("selected",f"{blue}")])
           
#WIDGETS  
#adding treeview
learner_tree_frame=ctk.CTkFrame(root,fg_color="transparent")

#learner treeview
#style 
style=tb.Style()
style.configure('Treeview',
               rowheight=20,
               font="times 13",
               fieldbackground="pink")
            #    background=f"{blue}")
style.configure("Treeview.Heading",foreground="blue",font="Times 13")
style.map('Treeview',background=[("selected",f"{blue}")])

learner_tree_scroll=Scrollbar(learner_tree_frame,orient=VERTICAL)
learner_tree_scroll.pack(side=RIGHT,fill=Y)
learner_tree_frame.place(x=30,y=100)
learner_tree=ttk.Treeview(learner_tree_frame,yscrollcommand=learner_tree_scroll.set,height=12,
selectmode="extended")
learner_tree_scroll.configure(command=learner_tree.yview)
learner_tree["columns"]=("no","adm","grade","name","paid","balance")
learner_tree.column("#0",width=0,stretch=NO)
learner_tree.column("no",width=40,anchor="center",minwidth="35")
learner_tree.column("adm",width=100,minwidth=80,anchor=CENTER)
learner_tree.column("grade",width=70,anchor="center",minwidth=50)
learner_tree.column("name",width=250,minwidth=180,anchor=W)
learner_tree.column("paid",width=80,minwidth=75,anchor=CENTER)
learner_tree.column("balance",width=80,minwidth=75,anchor=CENTER)
#headings
learner_tree.heading("#0",text="")
learner_tree.heading("no",text="#",anchor=CENTER)
learner_tree.heading("adm",text="ADM",anchor=CENTER)
learner_tree.heading("grade",text="GRADE",anchor=CENTER)
learner_tree.heading("name",text="NAME",anchor=CENTER)
learner_tree.heading("paid",text="PAID",anchor=CENTER)
learner_tree.heading("balance",text="BAL",anchor=CENTER)
learner_tree.pack()

#setting menubutton
menubutton=tb.Menubutton(root,text="Menu")
menubutton.place(x=30,y=10) 
menu=tb.Menu(menubutton)
menubutton["menu"]=menu 
admin_menu=tb.Menu(menu)
menu.add_cascade(menu=admin_menu,label="Admin")
admin_menu.add_command(label="Set Term",command=set_term_func)
admin_menu.add_separator()
admin_menu.add_command(label="Set School Name",command=school_name_func)
admin_menu.add_separator()
file_menu=tb.Menu(menu)
menu.add_cascade(menu=file_menu,label="Files")
file_menu.add_command(label="Archived Learners",command=display_archives)
gen_menu=tb.Menu(file_menu)
file_menu.add_cascade(menu=gen_menu,label="Generate files")
gen_menu.add_command(label="Remedial file-xlsx",command=generate_grade_reports,accelerator="ctrl+g")
gen_menu.add_command(label="Teacher attendance -xlsx",command=generate_teacher_attendance_report)
admin_menu.add_command(label="Register Persons",command=add_person_func)
menu.add_command(label="Learner Payment History",command=lambda:payment_history_func(None),accelerator="ctrl+H")
menu.add_separator()
# menu.add_command(label=" Display Tr. Attnd. History") #on the monitor table
menu.add_command(label="Monitor Tr. Attendance",command=lambda:teacher_win_func(None),accelerator="ctrl+A")
# menu.add_separator()
# menu.add_command(label="Attendance Archive", command=disp_attend_archive_func)

#within the table display teachers, their total_sessions, and total(already paid)
#popup menu
#popup func

learner_menu=Menu(root,tearoff=0)
learner_menu=Menu(learner_menu,tearoff=0)
learner_menu.add_command(label="Delete Learner",command=delete_learner)
learner_menu.add_separator()
learner_menu.add_command(label="Display Teachers",command=lambda:teacher_win_func(None))
learner_menu.add_separator()
learner_menu.add_command(label="Attendance History",command=tr_attendance_func)
learner_menu.add_separator()
learner_menu.add_command(label="Payment History",command=payment_history_func)

learner_search_entry=ctk.CTkEntry(root,width=150,font=("helvetica",16),
placeholder_text="Search")
learner_search_entry.place(x=152,y=46)
search_button=ctk.CTkButton(root,text="Search",width=50,command=search_func)
search_button.place(x=311,y=46)
weeks=["One","Two","Three","Four","Five","Six","Seven",
       "Eight","Nine","Ten","Eleven","Twelve","Thirteen",
       "Fourteen"]
week_label=ctk.CTkLabel(root,text="Week",font=("Helvetica",16))
week_label.place(x=390,y=10)
week_combo=ctk.CTkComboBox(root,height=30,width=120,font=("Helvetica",16),
                                      border_color=blue,button_color=blue,
                                      values=(weeks),state="readonly",command=set_default_week)
week_combo.place(x=440,y=10)
search_label=ctk.CTkLabel(root,text="")
search_label.place(x=376,y=46)
search_by_label=ctk.CTkLabel(root,text="Search By")

search_by_combo=ctk.CTkComboBox(root,height=30,width=150,font=("Helvetica",16),
                                      border_color=blue,button_color=blue,
                                 values=("Search By","Adm No","Name"),state="readonly")
search_by_combo.place(x=152,y=10)
search_by_combo.set("Search by")
disp_label=ctk.CTkLabel(root,text="")
disp_label.place(x=286,y=545)
toggle_switch=ctk.CTkSwitch(root,text="Switch Mode",command=toggle)
toggle_switch.place(x=47,y=571)
#learner frame widgets frame
lnr_tree_w_frame=ctk.CTkFrame(root,fg_color="transparent")
lnr_tree_w_frame.place(x=40,y=416)

amount_label=ctk.CTkLabel(lnr_tree_w_frame,text="Amount",font=("helvetica",16))
amount_label.grid(row=0,column=0,sticky=W)
amount_entry=ctk.CTkEntry(lnr_tree_w_frame,width=150,font=("helvetica",16))
amount_entry.grid(row=0,column=1,padx=10)
From_label=ctk.CTkLabel(lnr_tree_w_frame,text="Received From",font=("helvetica",16))
From_label.grid(row=1,column=0,sticky=W,pady=10)
From_entry=ctk.CTkEntry(lnr_tree_w_frame,width=150,font=("helvetica",16))
From_entry.grid(row=1,column=1,padx=10,pady=10)
grade=grades_list()
disp_bal=ctk.CTkLabel(lnr_tree_w_frame,font=("Helvetica",16),text="Grade")
disp_bal.grid(row=0,column=3,padx=10)
disp_bal_combo=ctk.CTkComboBox(lnr_tree_w_frame,height=30,width=190,font=("Helvetica",16),
                                      border_color=blue,button_color=blue,
                                      values=(grade),state="readonly",command=display_learners)
disp_bal_combo.set(grade[0])

disp_bal_combo.grid(row=0,column=4)
delete_trans_button=ctk.CTkButton(lnr_tree_w_frame,text="Delete Pay",width=100
                            ,command=delete_transaction)
delete_trans_button.grid(row=1,column=4)
# disp_bal_combo.set("Display Balances")
submit_button=ctk.CTkButton(lnr_tree_w_frame,text="Submit Pay",width=80
                            ,command=make_payment)
submit_button.grid(row=1,columnspan=3,column=3)
submit_button.grid(row=2,column=1)
#calling functions
retrieve_week()
display_learners()
#bindings
# root.bind("<Button-1>",placer)
root.bind("<Button-3>",binding)
# fee_tree_scroll.bind("<MouseWheel>")
#binding the update button
root.bind("<Return>",make_payment)
root.bind("<Control-t>",teacher_win_func)
root.bind("<Control-T>",teacher_win_func)
root.bind("<Control-h>",payment_history_func)
root.bind("<Control-H>",payment_history_func)
root.bind("<Control-a>",tr_attendance_func)
root.bind("<Control-A>",tr_attendance_func)
root.bind("<Control-G>",generate_records)
root.bind("<Control-g>",generate_records)
root.bind("<Control-C>",disp_attend_archive_func)
root.bind("<Control-c>",disp_attend_archive_func)
#saving learners and teachers records
# root.bind("<Control-z>",add_learner)
# root.bind("<Control-Z>",add_learner)
# root.bind("<Control-x>",add_teachers)
# root.bind("<Control-X>",add_teachers)
# #closing app
def confirm_close():
    root.bell()
    res=messagebox.askyesno("Remedial App","Are you sure you want to close\nRemedial App?")
    if res:
        root.destroy()
      


root.protocol("WM_DELETE_WINDOW",confirm_close)
root.mainloop()
