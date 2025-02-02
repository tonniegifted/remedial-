from tkinter import*
from tkinter import ttk
from tkinter import messagebox
import mysql.connector
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font
from decimal import Decimal


#connecting to database
my_db=mysql.connector.connect(host="localhost",
                              user="root",
                              password="print()",
                              database="student")

#creating cursor
cur=my_db.cursor()

root=Tk()

width=root.winfo_width()
height=root.winfo_height()
root.title("Remedial App")
path="D:/Tonniegifted/Remedial App/Resources/remedial convert.ico"
# root.iconbitmap(path)
root.geometry(f"{width}x{height}+0+0")  
# root.geometry("800x600+0+0")

def grade_balance(e):
    try:
        clear_tree()
        # global fee_tree
        grade=["Seven","Eight","Nine"]
        if disp_bal_combo=="Display balances by Grade":
            messagebox.showwarning("Remedial App","Select Grade to display")
        elif disp_bal_combo.get()=="Grade Seven":
            grd=grade[0]
        elif disp_bal_combo.get()=="Grade Eight":
            grd=grade[1]
        elif disp_bal_combo.get()=="Grade Nine":
            grd=grade[2]
        else:
            disp_bal_combo=="Display balances by Grade"
            messagebox.showwarning("Remedial App","Select Grade to display")
        cur.execute("""SELECT l.grade,l.learner_id,l.first,l.second,l.surname,
                    t.amount_paid,t.balance FROM learner l JOIN 
                    transaction t ON l.learner_id=t.learner_id WHERE l.grade=
                    %s ORDER BY t.amount_paid DESC""",(grd,))
        g_bal=cur.fetchall()
        if g_bal:
            for index,g_bal in enumerate(g_bal,start=1):
                fee_tree.insert("",END,values=(index,f"{g_bal[0]}".title(),g_bal[1],f"{g_bal[2]} {g_bal[3]} {g_bal[4]}".title(),g_bal[5],g_bal[6]))
        else:
            messagebox.showinfo("Remedial App",f"No payment records for Grade {grd}")
    except Exception:
        pass




def pay_token():
    try:
        teacher_id = adm_entry.get()
        teacher_id = int(teacher_id)
        
        # Get the active term ID
        cur.execute("SELECT term_id FROM term WHERE is_active=1")
        term_id = cur.fetchone()
        if term_id:
            active_term_id = term_id[0]
        else:
            messagebox.showerror("Remedial App", "No active term found.")
            return

        # Check if the teacher has any attendance records
        cur.execute("""
            SELECT COUNT(*)
            FROM teacher_attendance
            WHERE teacher_id = %s
        """, (teacher_id,))
        attendance_count = cur.fetchone()[0]
        
        if attendance_count == 0:
            messagebox.showerror("Remedial App", "Tokens cannot be paid for a teacher without\nattendance records.")
            return

        # Calculate the total session amount for the teacher's attendance records
        cur.execute("""
            SELECT SUM(session_amount)
            FROM teacher_attendance
            WHERE teacher_id = %s
        """, (teacher_id,))
        total_session_amount = cur.fetchone()[0]
        total_session_amount = Decimal(total_session_amount) if total_session_amount else Decimal(0)

        # Get the token amount paid
        token_paid = token_entry.get()
        try:
            token_paid = Decimal(token_paid)
        except ValueError:
            messagebox.showerror("Remedial App", "Invalid token amount. Please enter a valid number.")
            return

        # Validate that the token paid does not exceed the total session amount
        if token_paid > total_session_amount:
            messagebox.showerror("Remedial App", f"Token payment exceeds the total\nsession amount ({total_session_amount}).")
            return

        # Check the current token balance in the teacher_token table
        cur.execute("""
            SELECT token_paid 
            FROM teacher_token 
            WHERE teacher_id = %s AND term_id = %s
        """, (teacher_id, active_term_id))
        token = cur.fetchone()

        current_token_balance = Decimal(token[0]) if token else Decimal(0)

        # Ensure the entered token amount does not exceed the available balance
        required_balance = total_session_amount - current_token_balance
        if token_paid > required_balance:
            messagebox.showerror("Remedial App", f"The entered token amount exceeds the \nrequired balance of {required_balance}.")
            return

        # Add the token payment to the current balance
        new_token = current_token_balance + token_paid

        # Insert or update the teacher's token record in the teacher_token table
        cur.execute("""
            INSERT INTO teacher_token (teacher_id, term_id, token_paid)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE token_paid = VALUES(token_paid)
        """, (teacher_id, active_term_id, new_token))
        my_db.commit()

        # Retrieve attendance records in ascending order
        cur.execute("""
            SELECT teacher_attendance_id, session_amount
            FROM teacher_attendance
            WHERE teacher_id = %s
            ORDER BY teacher_attendance_id ASC
        """, (teacher_id,))
        attendance_records = cur.fetchall()

        payment_amount = new_token
        surplus_token = Decimal(0)

        # Process attendance records
        for record in attendance_records:
            attendance_id, session_amount = record
            session_amount = Decimal(session_amount)

            if payment_amount + surplus_token >= session_amount:
                # Archive attendance record
                cur.execute("""
                    INSERT INTO teacher_attendance_archive 
                        (teacher_id, week_number_id, session, grade, session_amount, subject, term_id)
                    SELECT teacher_id, week_number_id, session, grade, session_amount, subject, term_id
                    FROM teacher_attendance
                    WHERE teacher_attendance_id = %s
                """, (attendance_id,))

                # Delete the attendance record
                cur.execute("""
                    DELETE FROM teacher_attendance
                    WHERE teacher_attendance_id = %s
                """, (attendance_id,))

                # Deduct the session amount from payment
                payment_amount -= session_amount
            else:
                surplus_token = payment_amount + surplus_token
                payment_amount = Decimal(0)  # All tokens used
                break

        # Update the teacher's token balance with the remaining amount
        cur.execute("""
            UPDATE teacher_token
            SET token_paid = %s
            WHERE teacher_id = %s AND term_id = %s
        """, (payment_amount + surplus_token, teacher_id, active_term_id))

        # Commit changes
        my_db.commit()

        # Inform the user
        messagebox.showinfo("Remedial App", "Payment processing completed successfully.")
        token_entry.delete(0, END)

    except ValueError as ex:
        messagebox.showerror("Remedial App", f"Invalid input: {ex}\nPlease check your entries.")
    except Exception as ex:
        messagebox.showerror("Remedial App", f"An unexpected error occurred: {ex}")


def track_teacher_attendance():
    try:
        # Step 1: Get the active term ID
        cur.execute("""SELECT term_id FROM term WHERE is_active=1""")
        term = cur.fetchone()
        if not term:
            messagebox.showwarning("Remedial App", "No active term found. Please set an active term.")
            return
        term_id = term[0]

        # Step 2: Get the selected teacher ID from the treeview
        try:
            pos = fee_tree.selection()
            value = fee_tree.item(pos, "values")
            teacher_id = int(value[2])  # Extract and convert teacher_id
        except IndexError:
            messagebox.showerror("Remedial App", "Please select a teacher from the list.")
            return

        # Step 3: Get the active week number ID
        cur.execute("""SELECT week_number_id FROM week_number WHERE is_active=1 AND term_id=%s""", (term_id,))
        week_no = cur.fetchone()
        if not week_no:
            messagebox.showwarning("Remedial App", "No active week found. Please set an active week.")
            return
        week_number_id = week_no[0]

        # Step 4: Get the session and learning area
        session = tr_attend_combo.get()
        learning_area = subject_combo.get()
        if not session or not learning_area:
            messagebox.showerror("Remedial App", "Please select both the session and subject.")
            return

        # Step 5: Determine the session amount
        if session in ["Morning", "Evening"]:
            cur.execute("""SELECT weekday_pay FROM teacher_weekly_pay WHERE term_id=%s""", (term_id,))
        elif session == "Saturday":
            cur.execute("""SELECT weekend_pay FROM teacher_weekly_pay WHERE term_id=%s""", (term_id,))
        else:
            messagebox.showerror("Remedial App", "Invalid session type selected.")
            return

        pay = cur.fetchone()
        if not pay or pay[0] < 1:
            messagebox.showwarning("Remedial App", "Please set the weekly or weekend pay for the current term.")
            tr_lsn_pay()
            return
        session_amount = float(pay[0])

        # Step 6: Get the grade for the session
        tr_attend_grade = track_grade_combo.get()
        if not tr_attend_grade:
            messagebox.showerror("Remedial App", "Please select a grade.")
            return

        # # Step 7: Retrieve the teacher's total and deduct any paid amounts
        # cur.execute("""
        #     SELECT teacher_total FROM teacher_attendance 
        #     WHERE teacher_id=%s ORDER BY teacher_attendance_id DESC LIMIT 1
        # """, (teacher_id,))
        # prev = cur.fetchone()
        # prev_total = float(prev[0]) if prev else 0  # Use 0 if no previous total

        # # Check if the teacher has any payment recorded in teacher_token
        # cur.execute("""
        #     SELECT token_paid FROM teacher_token
        #     WHERE teacher_id=%s AND term_id=%s
        # """, (teacher_id, term_id))
        # payment = cur.fetchone()
        # payment_amount = float(payment[0]) if payment else 0

        # # Deduct the payment amount from the previous total
        # if payment_amount > 0:
        #     new_teacher_total = prev_total + session_amount - payment_amount

        #     # Clear the payment record to avoid repeated deductions
        #     cur.execute("""
        #         UPDATE teacher_token SET token_paid=0 
        #         WHERE teacher_id=%s AND term_id=%s
        #     """, (teacher_id, term_id))
        #     my_db.commit()
        # else:
        #     new_teacher_total = prev_total + session_amount

        # Step 8: Insert or update attendance record
        if teacher_id and week_number_id and session and tr_attend_grade and session_amount and learning_area:
            cur.execute("""
                INSERT  INTO teacher_attendance
                    (teacher_id, week_number_id, session, grade, session_amount,  subject, term_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (teacher_id, week_number_id, session, tr_attend_grade, session_amount, learning_area, term_id))
            my_db.commit()
            disp_label.config(text="Teacher attendance recorded successfully")
            disp_label.after(4000, lambda: disp_label.config(text=""))
        else:
            messagebox.showerror("Remedial App", "Submission failed. Please fill all required fields.")
            return

    except Exception as e:
        messagebox.showerror("Remedial App", f"Error: {e} Occurred")
    finally:
        track_grade_combo.set("")
        tr_attend_combo.set("")
        subject_combo.set("")

#setting week one as default in each new term
def week_one():
    cur.execute("UPDATE week_number SET is_active=0 WHERE is_active=1 ")
    cur.execute("""UPDATE week_number SET is_active=1 WHERE selected_week='One' AND 
                term_id=(SELECT term_id FROM term WHERE is_active=1)""")
    my_db.commit()
    cur.execute("SELECT selected_week FROM week_number WHERE is_active=1")
    week_number=cur.fetchone()[0]
    week_combo.set(week_number)
#setting default week
def set_default_week(e):
    selected_week=week_combo.get()
    #deactivating the rest of terms
    cur.execute("UPDATE week_number SET  is_active=0 WHERE is_active=1")
    my_db.commit()
    #updating selected term to active
    cur.execute("""UPDATE week_number SET is_active=1  WHERE selected_week=%s AND
                term_id=(SELECT term_id FROM term WHERE is_active=1)""",(selected_week,))
    my_db.commit()
    # clear_boxes()
#retrieving the set week
def retrieve_week():
      #retrieving selected term and setting it to default week
    cur.execute("""SELECT selected_week FROM week_number WHERE is_active=1
                AND term_id=(SELECT term_id FROM term WHERE is_active=1)""")
    default=cur.fetchone()
    if default:
       week_no=default[0]
       week_combo.set(week_no)
    else:
        week_combo.set(weeks[0] if weeks else '')
        # messagebox.showwarning("Remedial App","Set the current week first")
def seven_eight():
    root.bell()
    rep=messagebox.askyesno("Remedial App","Do you want to move learners?\nThis action can't be UNDONE")
    if rep:
        if grade_combo.get()=="Seven":
            cur.execute("UPDATE learner SET grade='Eight' WHERE grade= 'Seven'")
            my_db.commit()
            messagebox.showinfo("Remedial App","All Grade Seven learners have\nbeen moved to Grade Eight successfully")
        else:
            messagebox.showwarning("Remedial App","Ensure that you have selected\nGrade 'Seven' Before moving learners")
        disp_all_learners()
    else:
        pass
def eight_nine():
    root.bell()
    rep=messagebox.askyesno("Remedial App","Do you want to move learners?\nThis action can't be UNDONE")
    if rep:
        if grade_combo.get()=="Eight":
            cur.execute("UPDATE learner SET grade='Nine' WHERE grade ='Eight'")
            my_db.commit()
            messagebox.showinfo("Remedial App","All Grade Eight learners have\nbeen moved to Grade Nine successfully")
        else:
            messagebox.showwarning("Remedial App","Ensure that you have selected\nGrade 'Eight' Before moving learners")
        disp_all_learners()
    else:
        pass
def nine_archive():
    root.bell()
    rep=messagebox.askyesno("Remedial App","Do you want to move learners?\nThis action can't be UNDONE")
    if rep:
        if grade_combo.get()=="Nine":
            item=[]
            cur.execute("""SELECT * FROM learner WHERE grade='Nine'""")
            items=cur.fetchall()
            for i in items:
                item.append(i)
            cur.executemany("""INSERT INTO archive(learner_id,first,second,surname,grade)VALUES
                            (%s,%s,%s,%s,%s)""",(item))
            cur.execute("""DELETE FROM learner WHERE grade='Nine'""")
            my_db.commit()
            messagebox.showinfo("Remedial App","All Grade Nine learners have\nbeen moved to archives successfully")
        else:
            messagebox.showwarning("Remedial App","Ensure that you have selected\nGrade 'Nine' Before moving learners")
            archive_function()
            display_archives()
    else:
        pass
def delete_archive():
    global archive_label
    resp=messagebox.askyesno("Archives","Are you sure you want to delete learner\narchive records?")
    if resp:
        cur.execute("DELETE FROM archive")
        my_db.commit()
        disp_label.config("Remedial App","Learner Archive records Deleted\nsuccessfully")
        # archive_function()
        # display_archives()
    else:
        pass
#displaying archived learnersi
def display_archives():
    global archive_tree,archive_window
    archive_function()
    cur.execute("SELECT * FROM archive")
    item=cur.fetchall()
    if item:
            for index,item in enumerate(item,start=1):
                time=item[5]
                formatted_time=time.strftime("%d/%m/%Y")
               

                full_name=f"{item[1]} {item[2]} {item[3]}".title()
                archive_tree.insert("",END,values=(index,full_name,formatted_time))
    
    else:
        messagebox.showinfo("Archives","No records in Archives")
        archive_window.destroy()
    
#new window
def archive_function():
    # global search_entry2,archive_tree,disp_label2
    global archive_tree,archive_label,archive_window
    archive_window=Toplevel(root)
    archive_window.geometry("600x450+300+100")
    archive_window.title("Transaction History")
    #style 
    style=ttk.Style()
    style.configure('Treeview',
                rowheight=25,
                font="times 14")
    style.configure("Treeview.Heading",foreground="blue",font="Times 12 ")
    style.map('Treeview',background=[("selected","blue")])
    #fee collection treeview
    archive_tree=Frame(archive_window)
    archive_tree_scroll=Scrollbar(archive_tree,orient=VERTICAL)
    archive_tree_scroll.pack(side=RIGHT,fill=Y)
    archive_tree.pack()
    archive_tree=ttk.Treeview(archive_tree,yscrollcommand=archive_tree_scroll.set,height=12)
    archive_tree_scroll.configure(command=archive_tree.yview)
    #defining columns
    archive_tree["columns"]=('s/no','name','archived')
    archive_tree.column("#0",width=0,stretch=NO)
    archive_tree.column("s/no",width=40,anchor="center",minwidth=35)
    archive_tree.column("name",width=270,minwidth=180,anchor=W)
    archive_tree.column("archived",width=180,minwidth=180,anchor=W)
    #headings
    archive_tree.heading("#0",text="")
    archive_tree.heading("s/no",text="#",anchor=CENTER)
    archive_tree.heading("name",text="NAME",anchor=CENTER)
    archive_tree.heading("archived",text="DATE ARCHIVED",anchor=CENTER)
   
    #fee collection frame display
    archive_tree.pack(fill=BOTH,expand=True)
    # Restore/empty archive
    archive_del_button=Button(archive_window,text="Delete",font="times 11",command=delete_archive)
    archive_del_button.place(x=30,y=340)
    archive_label=Label(archive_window)
    archive_label.place(x=156,y=340)
#deleting selected teacher transaction records
def delete_attend_record():
    global tr_attend_tree,tr_attend_win
    try:
        res=messagebox.askyesno("Teacher Attendance History","Are you sure you want to delete\nThe last entered attendance record?")
        if res:
            # pos=tr_attend_tree.selection()
            # value=tr_attend_tree.item(pos,"values")
            # teacher_id=value[0]
            cur.execute("""SELECT teacher_attendance_id FROM teacher_attendance ORDER BY
                        record_date DESC LIMIT 1""")
            tr_attend_id=cur.fetchone()[0]
            # deleting last entered record by tr_id incase of a mistake
        cur.execute("""DELETE FROM teacher_attendance WHERE teacher_attendance_id=%s""",
                    (tr_attend_id,))
        my_db.commit()
    # tr_attend_tree.delete(value)
        messagebox.showinfo("Teacher Attendance History","Recent teacher attendance record deleted\nsuccessfully")
    except:
        messagebox.showerror("Teacher Attendance History","Unexpected error just occurred")
    # cur.execute("DELETE FROM teacher_attendance")
    finally:
        tr_attend_win.destroy()
#teacher attendance
def tr_attendance_func(e=None):
    global tr_attend_tree,tr_attend_search_entry,tr_attend_disp_label,tr_attend_win
    tr_attend_win=Toplevel(root)
    tr_attend_win.geometry("910x450+300+100")
    tr_attend_win.title("Teacher Attendance History")
    #style 
    style=ttk.Style()
    style.configure('Treeview',
                rowheight=25,
                font="times 14")
    style.configure("Treeview.Heading",foreground="blue",font="Times 12 ")
    style.map('Treeview',background=[("selected","blue")])
    #fee collection treeview
    tr_attend_frame=Frame(tr_attend_win)
    tr_attend_tree_scroll=Scrollbar(tr_attend_frame,orient=VERTICAL)
    tr_attend_tree_scroll.pack(side=RIGHT,fill=Y)
    tr_attend_frame.pack()
    tr_attend_tree=ttk.Treeview(tr_attend_frame,yscrollcommand=fee_tree_scroll.set,height=12,
                                )
    tr_attend_tree_scroll.configure(command=tr_attend_tree.yview)
    #defining columns
    tr_attend_tree["columns"]=('s/no','name','grade','subject','token','session','week','date')
    tr_attend_tree.column("#0",width=0,stretch=NO)
    tr_attend_tree.column("s/no",width=40,anchor="center",minwidth=35)
    tr_attend_tree.column("name",width=150,minwidth=150,anchor=W)
    tr_attend_tree.column("grade",width=80,minwidth=75,anchor=CENTER)
    tr_attend_tree.column("subject",width=60,minwidth=50,anchor=CENTER)
    tr_attend_tree.column("token",width=70,minwidth=50,anchor=CENTER)
    tr_attend_tree.column("session",width=100,minwidth=90,anchor=W)
    tr_attend_tree.column("week",width=150,minwidth=150,anchor=CENTER)
    tr_attend_tree.column("date",width=150,minwidth=150,anchor=W)
    #headings
    tr_attend_tree.heading("#0",text="")
    tr_attend_tree.heading("s/no",text="#",anchor=CENTER)
    tr_attend_tree.heading("name",text="NAME",anchor=CENTER)
    tr_attend_tree.heading("grade",text="GRADE",anchor=CENTER)
    tr_attend_tree.heading("subject",text="SUBJ",anchor=CENTER)
    tr_attend_tree.heading("token",text="TOKEN",anchor=CENTER)
    tr_attend_tree.heading("session",text="SESSION",anchor=CENTER)
    tr_attend_tree.heading("week",text="WEEK",anchor=CENTER)
    tr_attend_tree.heading("date",text="DATE",anchor=CENTER)
    #fee collection frame display
    tr_attend_tree.pack(fill=BOTH,expand=True,pady=50)
    #widget camera
    #search2
    # tr_attend_search_button=Button(tr_attend_win,text="Search",font="times 11",command=lambda: disp_attendance_history(None))
    tr_attend_search_button=Button(tr_attend_win,text="Search",command=disp_attendance_history,
                                                    width=10)
    # tr_attend_search_button.place(x=30,y=340)
    tr_attend_search_button.place(x=190,y=10)
    tr_attend_search_entry=Entry(tr_attend_win,width=80)
    # tr_attend_search_entry.place(x=90,y=340)
    tr_attend_search_entry.place(x=100,y=10)
  
    #button to delete teacher attendance per id
    tr_attend_delete_button=Button(tr_attend_win,text="Delete",font="times 11",command=delete_attend_record)
    tr_attend_delete_button.place(x=30,y=380)
    #display what should be searched
    tr_attend_search_entry.insert(0,"Tr.No")
    tr_attend_search_entry.bind("<FocusIn>",lambda e:tr_attend_search_entry.delete(0,END))
    #display_label2
    tr_attend_disp_label=Label(tr_attend_win,font="times 13")
    tr_attend_disp_label.place(x=250,y=10)
    #attempt to display all available records
    cur.execute("""SELECT t.first,t.second, a.grade,a.session,
            a.record_date,w.selected_week,a.session_amount,a.subject
            FROM teacher_attendance a JOIN teacher t ON t.teacher_id=a.teacher_id
            JOIN week_number w ON w.week_number_id=a.week_number_id JOIN term tm
            ON a.term_id=tm.term_id
            WHERE a.term_id=(SELECT term_id FROM term WHERE is_active=1)""")
    items=cur.fetchall()
    if items:
            for index,items in enumerate(items,start=1):
                time=items[4]
                formatted_time=time.strftime("%d-%a-%m-%Y")

                full_name=f"{items[0]} {items[1]}".title()
                tr_attend_tree.insert("",END,values=(index,full_name,items[2],items[7],items[6],items[3],items[5],formatted_time))
    
    else:
        messagebox.showinfo("Transaction history","No Teacher Attendance Records")
        tr_attend_win.destroy()   

def disp_attendance_history():  
    global  tr_attend_tree,tr_attend_search_entry,tr_attend_disp_label
    tr_attend_tree.delete(*tr_attend_tree.get_children()) 
    try:       
        teacher_id=tr_attend_search_entry.get()
        teacher_id=int(teacher_id)
    #displaying teacher attendance
        cur.execute("""SELECT t.first,t.second, a.grade,a.session,
            a.record_date,w.selected_week,a.session_amount,a.subject
            FROM teacher_attendance a JOIN teacher t ON t.teacher_id=a.teacher_id
            JOIN week_number w ON w.week_number_id=a.week_number_id JOIN term tm
            ON a.term_id=tm.term_id
            WHERE a.teacher_id=%s AND a.term_id=(SELECT term_id FROM term WHERE is_active=1)""",(teacher_id,))
        items=cur.fetchall()
        if items:
            for index,items in enumerate(items,start=1):
                time=items[4]
                formatted_time=time.strftime("%d-%a-%m-%Y")

                full_name=f"{items[0]} {items[1]}".title()
                tr_attend_tree.insert("",END,values=(index,full_name,items[2],items[7],items[6],items[3],items[5],formatted_time))
            cur.execute("""SELECT token_paid FROM teacher_token WHERE teacher_id
                        =%s""",(teacher_id,))
            tr_paid=cur.fetchone()
            if tr_paid:
                teacher_paid=tr_paid[0]
                teacher_paid=int(teacher_paid)
            else:
                teacher_paid=0
            cur.execute("""SELECT SUM(session_amount) FROM teacher_attendance
                        WHERE teacher_id=%s""",(teacher_id,))
            tot=cur.fetchone()
            if tot:
                total=int(tot[0])
                balance=(total-teacher_paid)
                tr_attend_disp_label.config(text=f"Balance= {balance}")
                # tr_attend_disp_label.after(4000,lambda: tr_attend_disp_label.config(text=""))
        else:
            tr_attend_disp_label.config(text=f"Record with Tr.No {tr_attend_search_entry.get()} not found")
            tr_attend_disp_label.after(4000,lambda: tr_attend_disp_label.config(text=""))
        
        tr_attend_search_entry.delete(0,END)
    except:
            tr_attend_disp_label.config(text="Search box cannot be blank")
            tr_attend_disp_label.after(4000,lambda:tr_attend_disp_label.config(text=""))
#teacher attendance
def tr_attendance_archive_func(e=None):
    global tr_attend_tree2,tr_attend_search_entry2,tr_attend_disp_label2,tr_attend_win2,tr_attend_delete_button2
    tr_attend_win2=Toplevel(root)
    tr_attend_win2.geometry("910x450+300+100")
    tr_attend_win2.title("Teacher Attendance Archives")
    #style 
    style=ttk.Style()
    style.configure('Treeview',
                rowheight=25,
                font="times 14")
    style.configure("Treeview.Heading",foreground="blue",font="Times 12 ")
    style.map('Treeview',background=[("selected","blue")])
    #fee collection treeview
    tr_attend_frame2=Frame(tr_attend_win2)
    tr_attend_tree_scroll2=Scrollbar(tr_attend_frame2,orient=VERTICAL)
    tr_attend_tree_scroll2.pack(side=RIGHT,fill=Y)
    tr_attend_frame2.pack()
    tr_attend_tree2=ttk.Treeview(tr_attend_frame2,yscrollcommand=tr_attend_tree_scroll2.set,height=12)
    tr_attend_tree_scroll2.configure(command=tr_attend_tree2.yview)
    #defining columns
    tr_attend_tree2["columns"]=('s/no','name','grade','subject','token','session','week','date')
    tr_attend_tree2.column("#0",width=0,stretch=NO)
    tr_attend_tree2.column("s/no",width=40,anchor="center",minwidth=35)
    tr_attend_tree2.column("name",width=150,minwidth=150,anchor=W)
    tr_attend_tree2.column("grade",width=80,minwidth=75,anchor=CENTER)
    tr_attend_tree2.column("subject",width=60,minwidth=50,anchor=CENTER)
    tr_attend_tree2.column("token",width=70,minwidth=50,anchor=CENTER)
    tr_attend_tree2.column("session",width=100,minwidth=90,anchor=W)
    tr_attend_tree2.column("week",width=150,minwidth=150,anchor=CENTER)
    tr_attend_tree2.column("date",width=150,minwidth=150,anchor=W)
    #headings
    tr_attend_tree2.heading("#0",text="")
    tr_attend_tree2.heading("s/no",text="#",anchor=CENTER)
    tr_attend_tree2.heading("name",text="NAME",anchor=CENTER)
    tr_attend_tree2.heading("grade",text="GRADE",anchor=CENTER)
    tr_attend_tree2.heading("subject",text="SUBJ",anchor=CENTER)
    tr_attend_tree2.heading("token",text="TOKEN",anchor=CENTER)
    tr_attend_tree2.heading("session",text="SESSION",anchor=CENTER)
    tr_attend_tree2.heading("week",text="WEEK",anchor=CENTER)
    tr_attend_tree2.heading("date",text="DATE",anchor=CENTER)
    #fee collection frame display
    tr_attend_tree2.pack(fill=BOTH,expand=True)
    #widget camera
    #search2
    # tr_attend_search_button=Button(tr_attend_win,text="Search",font="times 11",command=lambda: disp_attendance_history(None))
    tr_attend_search_button2=Button(tr_attend_win2,text="Search",font="times 11",command=disp_teacher_attend_archive)
    tr_attend_search_button2.place(x=30,y=340)
    tr_attend_search_entry2=Entry(tr_attend_win2,font="Helvetica 14",width=6)
    tr_attend_search_entry2.place(x=90,y=340)
    tr_attend_delete_button2=Button(tr_attend_win2,text="Delete",font="times 11",command=delete_tr_archive)
    tr_attend_delete_button2.place(x=30,y=380)
    #display what should be searched
    tr_attend_search_entry2.insert(0,"Tr.No")
    tr_attend_search_entry2.bind("<FocusIn>",lambda e:tr_attend_search_entry2.delete(0,END))
    #display_label2
    tr_attend_disp_label2=Label(tr_attend_win2)
    tr_attend_disp_label2.place(x=156,y=340)
    #attempt to display all available records
    cur.execute("""SELECT t.first,t.second, ar.grade,ar.session,
            ar.record_date,w.selected_week,ar.session_amount,ar.subject
            FROM teacher_attendance_archive ar JOIN teacher t ON t.teacher_id=ar.teacher_id
            JOIN week_number w ON w.week_number_id=ar.week_number_id JOIN term tm
            ON ar.term_id=tm.term_id
            WHERE ar.term_id=(SELECT term_id FROM term WHERE is_active=1)""")
    items=cur.fetchall()
    if items:
            for index,items in enumerate(items,start=1):
                time=items[4]
                formatted_time=time.strftime("%d-%a-%m-%Y")

                full_name=f"{items[0]} {items[1]}".title()
                tr_attend_tree2.insert("",END,values=(index,full_name,items[2],items[7],items[6],items[3],items[5],formatted_time))
    
    else:
        messagebox.showinfo("Attendance Archives","No Teacher Attendance archive records")
        tr_attend_win2.destroy()

def delete_tr_archive():
    global tr_attend_win2
    res=messagebox.askyesno("Teacher Attendance Archives","Are sure you want to delete all records\nThis action cannot be undone")
    if res:
        cur.execute("DELETE FROM teacher_attendance_archive")
        my_db.commit()
        messagebox.showinfo("Teacher Attendance Archives","Teacher attendance Archives deleted successfuly")
        
    else:
        tr_attend_win2.destroy()
#display archived teacher attendance records
def disp_teacher_attend_archive():  
    global  tr_attend_tree2,tr_attend_search_entry2,tr_attend_disp_label2
    tr_attend_tree2.delete(*tr_attend_tree2.get_children()) 
    try:       
        teacher_id=tr_attend_search_entry2.get()
        teacher_id=int(teacher_id)
    #displaying teacher attendance
        cur.execute("""SELECT t.first,t.second, ar.grade,ar.session,
            ar.record_date,w.selected_week,ar.session_amount,ar.subject
            FROM teacher_attendance_archive ar JOIN teacher t ON t.teacher_id=ar.teacher_id
            JOIN week_number w ON w.week_number_id=ar.week_number_id JOIN term tm
            ON ar.term_id=tm.term_id
            WHERE ar.teacher_id=%s AND ar.term_id=(SELECT term_id FROM term WHERE is_active=1)""",(teacher_id,))
        items=cur.fetchall()
        if items:
            for index,items in enumerate(items,start=1):
                time=items[4]
                formatted_time=time.strftime("%d-%a-%m-%Y")

                full_name=f"{items[0]} {items[1]}".title()
                tr_attend_tree2.insert("",END,values=(index,full_name,items[2],items[7],items[6],items[3],items[5],formatted_time))
            cur.execute("""SELECT SUM(session_amount) FROM teacher_attendance_archive
                        WHERE teacher_id=%s""",(teacher_id,))
            am=cur.fetchone()
            if am:
                amount_paid=am[0]
                tr_attend_disp_label2.config(text=f"Total Paid= {amount_paid}",font=
                                             "times 14")
            
        else:
            tr_attend_disp_label2.config(text=f"Record with Tr.No {tr_attend_search_entry2.get()} not found")
            tr_attend_disp_label2.after(4000,lambda: tr_attend_disp_label2.config(text=""))
    
        tr_attend_search_entry2.delete(0,END)
    except:
            tr_attend_disp_label2.config(text="Search box cannot be blank")
            tr_attend_disp_label2.after(4000,lambda:tr_attend_disp_label.config(text=""))

# def display_teacher_total():
#     try:
#         # Clear the tree before inserting new data
#         fee_tree.delete(*fee_tree.get_children())
        
#         # Step 1: Get all unique teacher IDs
#         cur.execute("SELECT DISTINCT teacher_id FROM teacher_token")
#         teacher_ids = cur.fetchall()
        
#         # Step 2: Loop through each teacher_id, retrieve the token_paid, and add numbering
#         for index, teacher in enumerate(teacher_ids, start=1):  # Use enumerate to add numbering
#             teacher_id = teacher[0]  # Extract teacher_id from the tuple
            
#             # Retrieve the `token_paid` for the current teacher
#             cur.execute("""
#                 SELECT token_paid 
#                 FROM teacher_token 
#                 WHERE teacher_id = %s
#             """, (teacher_id,))
#             token_data = cur.fetchone()
            
#             # Check if a token_paid value exists
#             if token_data:
#                 token_paid = token_data[0]
#             else:
#                 token_paid = 0  # Default to 0 if no token_paid found
            
#             # Step 3: Get the teacher details (first name, second name)
#             cur.execute("""
#                 SELECT t.teacher_id, t.first, t.second 
#                 FROM teacher t 
#                 WHERE t.teacher_id = %s
#             """, (teacher_id,))
#             teacher_details = cur.fetchone()
            
#             if teacher_details:
#                 teacher_id, first_name, second_name = teacher_details
#                 # Insert data into the Treeview with numbering
#                 fee_tree.insert(
#                     "", 
#                     "end", 
#                     values=(index,"x", teacher_id, f"{first_name} {second_name}".title(), f"- {token_paid}","0")
#                 )
#     except Exception as e:
#         messagebox.showerror("Teacher Attendance History",f"Error: {e}")

def display_teacher_total():
    try:
        # Clear the tree before inserting new data
        fee_tree.delete(*fee_tree.get_children())
        
        # Step 1: Get all unique teacher IDs
        cur.execute("SELECT DISTINCT teacher_id FROM teacher_token")
        teacher_ids = cur.fetchall()
        
        # Check if there are any teachers
        if not teacher_ids:
            messagebox.showinfo("Teacher Attendance History", "No teacher records found.")
            display_tr()  # Call display_tr() when no teachers are found
            return

        # Step 2: Loop through each teacher_id, retrieve the token_paid, and add numbering
        for index, teacher in enumerate(teacher_ids, start=1):  # Use enumerate to add numbering
            teacher_id = teacher[0]  # Extract teacher_id from the tuple
            
            # Retrieve the `token_paid` for the current teacher
            cur.execute("""
                SELECT token_paid 
                FROM teacher_token 
                WHERE teacher_id = %s
            """, (teacher_id,))
            token_data = cur.fetchone()
            
            # Check if a token_paid value exists
            if token_data:
                token_paid = token_data[0]
            else:
                token_paid = 0  # Default to 0 if no token_paid found
            
            # Step 3: Get the teacher details (first name, second name)
            cur.execute("""
                SELECT t.teacher_id, t.first, t.second 
                FROM teacher t 
                WHERE t.teacher_id = %s
            """, (teacher_id,))
            teacher_details = cur.fetchone()
            
            if teacher_details:
                teacher_id, first_name, second_name = teacher_details
                # Insert data into the Treeview with numbering
                fee_tree.insert(
                    "", 
                    "end", 
                    values=(index, "x", teacher_id, f"{first_name} {second_name}".title(), f"- {token_paid}", "0")
                )
        
        # If no data was inserted into the Treeview, display an info message
        if len(fee_tree.get_children()) == 0:
            messagebox.showinfo("Teacher Attendance History", "No teacher totals found.")
            display_tr()  # Call display_tr() after no records are found
        
    except Exception as e:
        messagebox.showerror("Teacher Attendance History", f"Error: {e}")  # Show any error that occurs



# new window
def top_window_func(e=None):
    global search_entry2,fee_tree2,disp_label2
    top_window=Toplevel(root)
    top_window.geometry("910x450+300+100")
    top_window.title("Transaction History")
    #style F
    style=ttk.Style()
    style.configure('Treeview',
                rowheight=25,
                font="times 14")
    style.configure("Treeview.Heading",foreground="blue",font="Times 12 ")
    style.map('Treeview',background=[("selected","blue")])
    #fee collection treeview
    fee_tree_frame2=Frame(top_window)
    fee_tree_scroll=Scrollbar(fee_tree_frame2,orient=VERTICAL)
    fee_tree_scroll.pack(side=RIGHT,fill=Y)
    fee_tree_frame2.pack()
    fee_tree2=ttk.Treeview(fee_tree_frame2,yscrollcommand=fee_tree_scroll.set,height=12)
    fee_tree_scroll.configure(command=fee_tree2.yview)
    #defining columns
    fee_tree2["columns"]=('s/no','name','paid','balance','time','comment')
    fee_tree2.column("#0",width=0,stretch=NO)
    fee_tree2.column("s/no",width=40,anchor="center",minwidth=35)
    fee_tree2.column("name",width=270,minwidth=180,anchor=W)
    fee_tree2.column("paid",width=80,minwidth=75,anchor=CENTER)
    fee_tree2.column("balance",width=80,minwidth=75,anchor=CENTER)
    fee_tree2.column("time",width=180,minwidth=180,anchor=W)
    fee_tree2.column("comment",width=150,minwidth=150,anchor=W)
    #headings
    fee_tree2.heading("#0",text="")
    fee_tree2.heading("s/no",text="#",anchor=CENTER)
    fee_tree2.heading("name",text="NAME",anchor=CENTER)
    fee_tree2.heading("paid",text="PAID",anchor=CENTER)
    fee_tree2.heading("balance",text="BAL",anchor=CENTER)
    fee_tree2.heading("time",text="TIME",anchor=CENTER)
    fee_tree2.heading("comment",text="COMMENT",anchor=CENTER)
    #fee collection frame display
    fee_tree2.pack(fill=BOTH,expand=True)
    #widget camera
    #search2
    search_button2=Button(top_window,text="Search",font="times 11",command=lambda:disp_trans_history(None))
    search_button2.place(x=30,y=340)
    search_entry2=Entry(top_window,font="Helvetica 14",width=6)
    search_entry2.place(x=90,y=340)
    #display what should be searched
    search_entry2.insert(0,"Adm")
    search_entry2.bind("<FocusIn>",lambda e:search_entry2.delete(0,END))
    #display_label2
    disp_label2=Label(top_window)
    disp_label2.place(x=156,y=340)
    #attempt to display all available records
    cur.execute("""SELECT l.learner_id,l.first,l.second,l.surname,th.amount,
                    th.balance,th.trans_time,th.comment FROM learner l JOIN transaction_history th ON
                    l.learner_id=th.learner_id WHERE th.term_id=(
                    SELECT term_id FROM term WHERE is_active=1) ORDER BY trans_time DESC""") 
    items=cur.fetchall()
    if items:
            for index,items in enumerate(items,start=1):
                time=items[6]
                formatted_time=time.strftime("%d/%m/%Y %I:%M:%p")

                full_name=f"{items[1]} {items[2]} {items[3]}".title()
                fee_tree2.insert("",END,values=(index,full_name,items[4],items[5],formatted_time,items[7]))
    
    else:
        messagebox.showinfo("Transaction history","No Transaction History Records")
        top_window.destroy()
#displaying transaction history
def disp_trans_history(e=None):  
    global search_entry2,fee_tree2,disp_label2
    fee_tree2.delete(*fee_tree2.get_children()) 
    try:       
        learner_id=search_entry2.get()
        learner_id=int(learner_id)
    #displaying transaction history
        cur.execute("""SELECT l.learner_id,l.first,l.second,l.surname,th.amount,
                    th.balance,th.trans_time,th.comment FROM learner l JOIN transaction_history th ON
                    l.learner_id=th.learner_id WHERE l.learner_id=%s AND th.term_id=(
                    SELECT term_id FROM term WHERE is_active=1) ORDER BY trans_time DESC""",(learner_id,) )
        items=cur.fetchall()
        if items:
            for index,items in enumerate(items,start=1):
                time=items[6]
                formatted_time=time.strftime("%d/%m/%Y %I:%M:%p")

                full_name=f"{items[1]} {items[2]} {items[3]}".title()
                fee_tree2.insert("",END,values=(index,full_name,items[4],items[5],formatted_time,items[7]))
        else:
            disp_label2.config(text=f"Record with Adm {search_entry2.get()} not found")
            disp_label2.after(4000,lambda:disp_label2.config(text=""))
        
        search_entry2.delete(0,END)
    except:
            disp_label2.config(text="Search box cannot be blank")
            disp_label2.after(4000,lambda:disp_label2.config(text=""))
#generating all class lists
def generate_class_list():
    try:
        generate_g7()
        generate_g8()
        generate_g9()
        messagebox.showinfo("Remedial App","Class generate sucessfully")
    except:
        messagebox.showerror("Unknown error just occurred")
#generating all remedial fee/attendance lists
def remedial_records_to_excel(e=None):
    try:
        transaction_to_excel()
        attendance_archive_to_excel()
        attendance_to_excel()
        messagebox.showinfo("Remedial App","Remedial Records generated sucessfully")
    except:
        messagebox.showerror("Unknown error just occurred")
#generating excel files
def generate_g7():
    path="D:/Tonniegifted/Remedial App/G7 Class List.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Learners Remedial Balances" 
#generating remedial transaction history
    nested_list=[]
    cur.execute("""SELECT grade,learner_id, first,second,surname FROM learner
                WHERE grade='Seven' ORDER BY
                learner_id""")
    items=cur.fetchall()
    if items:
            for index,items in enumerate(items,start=1):
                full_name=f"{items[2]} {items[3]} {items[4]}".title()
                row=(index,items[0],items[1],full_name)
                nested_list.append(row)
            headings = ["#","GRADE","ADM","NAME"]
            for col_num, heading in enumerate(headings, start=1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = heading
                cell.font = Font(bold=True)  # Make the font bold
        #writing excel into sheet
            for item in nested_list:
                ws.append(item)
            # Autofit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter  # Get the column letter (e.g., 'A', 'B', ...)
                for cell in column:
                    if cell.value:
                        # Calculate the length of the cell value (converted to a string)
                        max_length = max(max_length, len(str(cell.value)))
                # Adjust column width slightly for better aesthetics
                ws.column_dimensions[column_letter].width = max_length + 2
            wb.save(path)
def generate_g8():
    path="D:/Tonniegifted/Remedial App/G8 Class List.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Learners Remedial Balances" 
#generating remedial transaction history
    nested_list=[]
    cur.execute("""SELECT grade,learner_id, first,second,surname FROM learner
                WHERE grade='Eight' ORDER BY
                learner_id""")
    items=cur.fetchall()
    if items:
            for index,items in enumerate(items,start=1):
                full_name=f"{items[2]} {items[3]} {items[4]}".title()
                row=(index,items[0],items[1],full_name)
                nested_list.append(row)
            headings = ["#","GRADE","ADM","NAME"]
            for col_num, heading in enumerate(headings, start=1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = heading
                cell.font = Font(bold=True)  # Make the font bold
        #writing excel into sheet
            for item in nested_list:
                ws.append(item)
            # Autofit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter  # Get the column letter (e.g., 'A', 'B', ...)
                for cell in column:
                    if cell.value:
                        # Calculate the length of the cell value (converted to a string)
                        max_length = max(max_length, len(str(cell.value)))
                # Adjust column width slightly for better aesthetics
                ws.column_dimensions[column_letter].width = max_length + 2
            wb.save(path)
def generate_g9():
    path="D:/Tonniegifted/Remedial App/G9 Class List.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "G9 Class List" 
#generating remedial transaction history
    nested_list=[]
    cur.execute("""SELECT grade,learner_id, first,second,surname FROM learner
                WHERE grade='Nine' ORDER BY
                learner_id""")
    items=cur.fetchall()
    if items:
            for index,items in enumerate(items,start=1):
                full_name=f"{items[2]} {items[3]} {items[4]}".title()
                row=(index,items[0],items[1],full_name)
                nested_list.append(row)
            headings = ["#","GRADE","ADM","NAME"]
            for col_num, heading in enumerate(headings, start=1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = heading
                cell.font = Font(bold=True)  # Make the font bold
        #writing excel into sheet
            for item in nested_list:
                ws.append(item)
            # Autofit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter  # Get the column letter (e.g., 'A', 'B', ...)
                for cell in column:
                    if cell.value:
                        # Calculate the length of the cell value (converted to a string)
                        max_length = max(max_length, len(str(cell.value)))
                # Adjust column width slightly for better aesthetics
                ws.column_dimensions[column_letter].width = max_length + 2
            wb.save(path)
#generating excel file for learner balances
def transaction_to_excel():
    path="D:/Tonniegifted/Remedial App/Remedial Balances.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Learners Remedial Balances" 
#generating remedial transaction history
    nested_list=[]
    cur.execute("""SELECT l.grade,l.learner_id,l.first,l.second,
                l.surname,t.amount_paid,t.balance,t.time_paid FROM learner l 
                JOIN transaction t  ON l.learner_id=t.learner_id 
                WHERE term_id=(SELECT term_id FROM term WHERE
                is_active=1)
                ORDER BY l.grade DESC""")
    items=cur.fetchall()
    if items:
            for index,items in enumerate(items,start=1):
                time=items[7]
                formatted_time=time.strftime("%d %b,%Y %I:%M:%p")

                full_name=f"{items[1]} {items[2]} {items[3]}".title()
                # fee_tree2.insert("",END,values=(index,full_name,items[4],items[5],formatted_time,items[7]))
                row=(index,items[1],full_name,items[0],items[5],items[6],formatted_time)
                nested_list.append(row)
            headings = ["#","ADM","NAME","GRADE","PAID","BALANCE","DATE/TIME"]
            for col_num, heading in enumerate(headings, start=1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = heading
                cell.font = Font(bold=True)  # Make the font bold
        #writing excel into sheet
            for item in nested_list:
                ws.append(item)
            # Autofit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter  # Get the column letter (e.g., 'A', 'B', ...)
                for cell in column:
                    if cell.value:
                        # Calculate the length of the cell value (converted to a string)
                        max_length = max(max_length, len(str(cell.value)))
                # Adjust column width slightly for better aesthetics
                ws.column_dimensions[column_letter].width = max_length + 2
            wb.save(path)
def attendance_to_excel():
    # headings = ["#","ADM","NAME","GRADE","PAID","BALANCE","DATE/TIME"]
    path="D:/Tonniegifted/Remedial App/Teacher Attendance Records.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Teacher Attendance Records" 
#generating remedial transaction history
    nested_list=[]
    cur.execute("""SELECT t.first,t.second, a.grade,a.session,
        a.record_date,w.selected_week,a.session_amount,a.subject
        FROM teacher_attendance a JOIN teacher t ON t.teacher_id=a.teacher_id
        JOIN week_number w ON w.week_number_id=a.week_number_id JOIN term tm
        ON a.term_id=tm.term_id
        WHERE a.term_id=(SELECT term_id FROM term WHERE is_active=1)""")
    items=cur.fetchall()
    if items:
            for index,items in enumerate(items,start=1):
                time=items[4]
                formatted_time=time.strftime("%d-%a-%m-%Y")

                full_name=f"{items[0]} {items[1]}".title()
                row=(index,full_name,items[2],items[7],items[6],items[3],items[5],formatted_time)
                nested_list.append(row)
            headings = ["#","NAME","GRADE","SUBJ","TOKEN","SESSION","WEEK","DATE"]
            for col_num, heading in enumerate(headings, start=1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = heading
                cell.font = Font(bold=True)  # Make the font bold
        #writing excel into sheet
            for item in nested_list:
                ws.append(item)
            # Autofit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter  # Get the column letter (e.g., 'A', 'B', ...)
                for cell in column:
                    if cell.value:
                        # Calculate the length of the cell value (converted to a string)
                        max_length = max(max_length, len(str(cell.value)))
                # Adjust column width slightly for better aesthetics
                ws.column_dimensions[column_letter].width = max_length + 2
            wb.save(path)
def attendance_archive_to_excel():
    # headings = ["#","ADM","NAME","GRADE","PAID","BALANCE","DATE/TIME"]
    path="D:/Tonniegifted/Remedial App/Teacher Attendance archive.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Teacher Attendance Archive" 
#generating remedial transaction history
    nested_list=[]
    cur.execute("""SELECT t.first,t.second, a.grade,a.session,
        a.record_date,w.selected_week,a.session_amount,a.subject
        FROM teacher_attendance_archive a JOIN teacher t ON t.teacher_id=a.teacher_id
        JOIN week_number w ON w.week_number_id=a.week_number_id JOIN term tm
        ON a.term_id=tm.term_id
        WHERE a.term_id=(SELECT term_id FROM term WHERE is_active=1)""")
    items=cur.fetchall()
    if items:
            for index,items in enumerate(items,start=1):
                time=items[4]
                formatted_time=time.strftime("%d-%a-%m-%Y")

                full_name=f"{items[0]} {items[1]}".title()
                row=(index,full_name,items[2],items[7],items[6],items[3],items[5],formatted_time)
                nested_list.append(row)
            headings = ["#","NAME","GRADE","SUBJ","TOKEN","SESSION","WEEK","DATE"]
            for col_num, heading in enumerate(headings, start=1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = heading
                cell.font = Font(bold=True)  # Make the font bold
        #writing excel into sheet
            for item in nested_list:
                ws.append(item)
            # Autofit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter  # Get the column letter (e.g., 'A', 'B', ...)
                for cell in column:
                    if cell.value:
                        # Calculate the length of the cell value (converted to a string)
                        max_length = max(max_length, len(str(cell.value)))
                # Adjust column width slightly for better aesthetics
                ws.column_dimensions[column_letter].width = max_length + 2
            wb.save(path)
#searching using adm number
def search_by_id():
    clear_tree()
    try:
        learner_id=search_entry.get()
        learner_id=int(learner_id)
        #checking learner payment status
        cur.execute("""SELECT l.grade,l.learner_id,l.first,l.second,l.surname,t.amount_paid,
                    t.balance FROM learner l JOIN transaction t  ON 
                    l.learner_id=t.learner_id
                    WHERE t.learner_id=%s AND term_id=(SELECT term_id FROM term
                    WHERE is_active=1)
                    """,(learner_id,))
        row=cur.fetchall()
        if row:
            for index,row in enumerate(row, start=1):
                    fee_tree.insert('',END,values=(index,f"{row[0]}".title(),row[1],f"{row[2]} {row[3]} {row[4]}".title(),row[5],row[6]))
    #if not learner payment record search in learners list
        else:
            cur.execute("SELECT * FROM learner WHERE learner_id=%s",(learner_id,))
            row=cur.fetchall()
            if row:
                for index,row in enumerate(row,start=1):
                    fee_tree.insert('',END,values=(index,f"{row[4]}".title(),row[0],f"{row[1]} {row[2]} {row[3]}".title()))
            else:
                search_label.config(text="Not Found")
                search_label.after(3000,lambda:search_label.config(text=""))
    except ValueError :
        messagebox.showwarning("Remedial","Enter Adm to search a learner")
    search_entry.delete(0,END)
#
#setting learners termly payable fee
def learner_term_pay():
    wk_label.place_forget()
    wk_entry.place_forget()
    wkend_label.place_forget()
    wkend_entry.place_forget()
    tr_lesson_pay_button.place_forget()
    tr_pay_submit_button.place_forget()
    
    lnr_amount_label.place (x=500,y=22)
    lnr_amount_entry.place (x=562,y=22)
    Grade_label.place(x=376,y=22) 
    grade_entry_combo.place(x=438,y=22)
    l_pay_submit_button.place(x=638,y=22)
    
def hiding_l_widgets():
    lnr_amount_label.place_forget()
    lnr_amount_entry.place_forget()
    Grade_label.place_forget()
    grade_entry_combo.place_forget()
    l_pay_submit_button.place_forget()
    tr_lesson_pay_button.place(x=386,y=22)
def hiding_tr_widgets():
    wk_label.place_forget()
    wk_entry.place_forget()
    wkend_label.place_forget()
    wkend_entry.place_forget()
    tr_pay_submit_button.place_forget()
    wk_entry.delete(0,END)
    wkend_entry.delete(0,END)
def submit_learner_fee():
    root.bell()
    response=messagebox.askokcancel("Remedial App","This will save changes")
    if response==1:
        try:
            amount=lnr_amount_entry.get()
            grade=grade_entry_combo.get()
            grade=int(grade)
            amount=float(amount)
            #accessing the term
            cur.execute("SELECT term_id FROM term  WHERE is_active=1")
            term=cur.fetchone()
            term_id=term[0]
                
            if amount  and grade and term_id:
                # logic.setting_learner_amount(t,amount,grade)
                cur.execute("""UPDATE termly_pay
                SET amount =CASE
                WHEN grade=%s AND term_id=%s THEN %s
                ELSE amount
                END
                WHERE term_id IN (%s)
                """,(grade,term_id,amount,term_id))
                my_db.commit()
            lnr_amount_entry.delete(0,END)
            grade_entry_combo.delete(0,END)
            hiding_l_widgets()
            disp_label.config(text=f"Fee payable for {grade} was set Successfully")
            disp_label.after(4000,lambda:disp_label.config(text=""))
        except ValueError:
            messagebox.showerror("Remedial App","Blank Fields cannot be set")
            hiding_l_widgets()
    else:
        pass
def tr_lsn_pay():
    wk_label.place(x=479,y=22)
    wk_entry.place(x=574,y=22)
    wkend_label.place(x=636,y=22)
    wkend_entry.place(x=730,y=22)
    tr_pay_submit_button.place(x=810,y=22)
    
def submit_tr_lsn_pay():
    root.bell()
    response=messagebox.askokcancel("Remedial App","This will save changes")
    if response==1:
        try:
            wk_pay=wk_entry.get()
            wk_pay=float(wk_pay)
            wkend_pay=wkend_entry.get()
            wkend_pay=float(wkend_pay)
    # ('teacher_weekly_pay_id', 'int', 'NO', 'PRI', None, 'auto_increment')
    # ('term_id', 'int', 'NO', 'MUL', None, '')
    # ('weekday_pay', 'decimal(10,0)', 'YES', '', '0', '')
    # ('weekend_pay', 'decimal(10,0)', 'YES', '', '0', '')
            cur.execute("""UPDATE teacher_weekly_pay SET weekday_pay=%s,weekend_pay=%s
                        WHERE term_id=(SELECT term_id FROM term WHERE is_active=1)""",
                        (wk_pay,wkend_pay))
            my_db.commit()
            disp_label.config(text="Weekly teacher token was set successfully")
            disp_label.after(4000,lambda:disp_label.config(text=""))
            hiding_tr_widgets()
        except ValueError:
            messagebox.showerror("Remedial App","Blank Fields cannot be set")
            hiding_tr_widgets()
    else:
        pass



#insert and update teachers
def add_teachers(e=None):
    try:
            teacher_id=adm_entry.get()
            teacher_id=int(teacher_id)
            first=first_entry.get()
            second=second_entry.get()
            surname=surname_entry.get()
            if teacher_id and first and second:
                root.bell()
                resp=messagebox.askyesno("Remedial App","Are sure you want to add to TEACHERS LIST?")
                if resp:
            #checking whether the entered adm exist 
                    cur.execute("""SELECT teacher_id FROM teacher WHERE 
                                teacher_id=%s""",(teacher_id,))
                    l_id=cur.fetchone()
                    l_id= l_id[0] if l_id else None
            #inserting learners if they do not exist
                    if not l_id:
                        cur.execute("""INSERT INTO teacher(teacher_id,first,second,surname
                                    )VALUES(%s,%s,%s,%s)""",(teacher_id,first,second,surname))
                        my_db.commit()
                        clear_boxes()
            #updating learners if they exist
                    else:
                        root.bell()
                        response=messagebox.askyesno("Remedial App",f"Tr.No. {teacher_id} exists do you want to overwrite?")
                        if response:
                            cur.execute("""UPDATE teacher SET first=%s,second=%s,surname=%s
                                        WHERE teacher_id=%s""",(first,second,surname,teacher_id))
                            my_db.commit()
                            clear_boxes()
                        else:
                            pass
                
                else:
                    pass    
            else:
                messagebox.showinfo("Remedial App","Blank Records can't be added")
                
            
    except:
        messagebox.showerror("Remedial App","An error occurred Check\nyour Values and try again")
    display_tr()
        
        
#selecting from fee tree
def populate(e):
    clear_boxes()
    try:
        position=fee_tree.selection()
        # for pos in  position:
        if position:
            pos=position[0]
            values=fee_tree.item(pos,"values")  
            adm_entry.insert(0,values[2])
            full_name=values[3].split()
            first_entry.insert(0,full_name[0])
            second_entry.insert(0,full_name[1])
            surname_entry.insert(0,full_name[2])
            if len(values[1])>=4:
                grade_combo.set(values[1])
            else:
                pass
    except IndexError :
            pass
#display all learners
def  disp_all_learners(e=None):
    clear_tree()
    cur.execute("SELECT * FROM learner  ORDER BY learner_id")
    row=cur.fetchall()
    if row:
        for index,row in enumerate(row,start=1):
            fee_tree.insert('',END,values=(index,f"{row[4]}".title(),row[0],f"{row[1]} {row[2]} {row[3]}".title()))
    else:
        messagebox.showinfo("Remedial App","Learners List is empty")
    

#add/update learners
def add_learner(e=None):
        try:
            learner_id=adm_entry.get()
            learner_id=int(learner_id)
            first=first_entry.get()
            second=second_entry.get()
            surname=surname_entry.get()
            grad=grade_combo.get()
            if learner_id and first and second and grad:
                root.bell()
                resp=messagebox.askyesno("Remedial App","Are sure you want to add to LEARNERS LIST?")
                if resp:
            #checking whether the entered adm exist 
                    cur.execute("""SELECT learner_id FROM learner WHERE 
                                learner_id=%s""",(learner_id,))
                    l_id=cur.fetchone()
                    l_id= l_id[0] if l_id else None
            #inserting learners if they do not exist
                    if not l_id:
                        cur.execute("""INSERT INTO learner(learner_id,first,second,surname,grade
                                    )VALUES(%s,%s,%s,%s,%s)""",(learner_id,first,second,surname,grad))
                        my_db.commit()
                        clear_boxes()
            #updating learners if they exist
                    else:
                        root.bell()
                        response=messagebox.askyesno("Remedial App",f"Adm No {learner_id} exits do you want to overwrite?")
                        if response:
                            cur.execute("""UPDATE learner SET first=%s,second=%s,surname=%s,grade=%s
                                        WHERE learner_id=%s""",(first,second,surname,grad,learner_id))
                            my_db.commit()
                            clear_boxes()
                            
                        else:
                            pass
                
                else:
                    pass    
            else:
                messagebox.showinfo("Remedial App","Blank Records can't be added")
                
            
        except:
            messagebox.showerror("Remedial App","An error occurred Check\nyour Values and try again")
        disp_all_learners()

#MAKE PAYMENT
# 
def make_payment(e):
    try:
        learner_id = adm_entry.get()
        learner_id = int(learner_id)
        amount_paid = amount_paid_entry.get()
        amount_paid = float(amount_paid)
        
        # Retrieve the current term's number
        cur.execute("SELECT term_number FROM term WHERE is_active = 1")
        term_no = cur.fetchone()
        if term_no:
            current_term_number = term_no[0]
        else:
            messagebox.showwarning("Remedial App", "No active term found. Please check term setup.")
            return
        
        # If the current term is term one, skip previous term balance checks
        if current_term_number == 1:
            prev_term_id = None  # No previous term exists for term one
        else:
            # Retrieve the previous term ID
            cur.execute("SELECT term_id FROM term WHERE term_number = %s", (current_term_number - 1,))
            prev_id = cur.fetchone()
            prev_term_id = prev_id[0] if prev_id else None
            
            if not prev_term_id:
                messagebox.showwarning("Remedial App", "Could not find previous term details. Please verify terms.")
                return
            
            # Check if the learner has any balance for the previous term
            cur.execute("""SELECT amount_paid, balance FROM transaction 
                            WHERE learner_id = %s AND term_id = %s""", (learner_id, prev_term_id))
            balance = cur.fetchone()
            
            if balance:
                prev_amount_paid = balance[0]  # Amount already paid for the previous term
                prev_balance = balance[1]     # Remaining balance for the previous term
            else:
                prev_amount_paid = 0
                prev_balance = 0  # No record means no payments made
            
            # If the learner has not fully cleared the previous term's balance, block payment
            if prev_balance > 0 or prev_amount_paid == 0:
                cur.execute("SELECT first, second, surname FROM learner WHERE learner_id = %s", (learner_id,))
                name = cur.fetchone()
                fullname = f"\n{name[0].title()} {name[1].title()} {name[2].title()}"
                messagebox.showwarning(
                    "Remedial App", 
                    f"{fullname} has not cleared\ntheir previous term fees."
                    f"Amount paid so\nfar is: KSH {prev_amount_paid:.2f}.")
                return
        
        # Retrieve termly pay for the learner's grade
        cur.execute("SELECT grade FROM learner WHERE learner_id = %s", (learner_id,))
        grade_data = cur.fetchone()
        if not grade_data:
            messagebox.showwarning("Remedial App", "Learner grade not found. Please verify learner details.")
            return
        
        grade = grade_data[0]
        grade = 7 if grade == "Seven" else 8 if grade == "Eight" else 9
        
        cur.execute("""SELECT amount FROM termly_pay WHERE grade = %s
                        AND term_id = (SELECT term_id FROM term WHERE is_active = 1)""", (grade,))
        amt = cur.fetchone()
        tot = float(amt[0]) if amt else 0
        
        if tot < 1:
            cur.execute("SELECT selected_term FROM term WHERE is_active = 1")
            is_active = cur.fetchone()[0]
            messagebox.showwarning("Remedial App", f"Set Fee payable for Grade\n{grade_combo.get()} {is_active}")
            learner_term_pay()
            return
        
        # Check for existing transaction records for the current term
        cur.execute("""SELECT amount_paid, balance FROM transaction 
                        WHERE learner_id = %s AND term_id = (SELECT term_id FROM term WHERE is_active = 1)""", (learner_id,))
        record = cur.fetchone()
        
        if record:
            new_amount = float(record[0]) + amount_paid
            new_balance = float(record[1]) - amount_paid
        else:
            new_amount = amount_paid
            new_balance = tot - amount_paid
        
        # Ensure the amount paid doesn't exceed the termly fee
        if new_amount > tot:
            messagebox.showwarning("Remedial App", "Amount Paid cannot exceed termly\npayable amount")
            return
        
        # Retrieve the current term ID
        cur.execute("SELECT term_id FROM term WHERE is_active = 1")
        active = cur.fetchone()[0]
        
        # Insert or update the transaction
        cur.execute("""INSERT INTO transaction(amount_paid, learner_id, term_id, balance)
                        VALUES (%s, %s, %s, %s)
                        ON DUPLICATE KEY UPDATE
                        amount_paid = VALUES(amount_paid),
                        balance = VALUES(balance)""", (new_amount, learner_id, active, new_balance))
        my_db.commit()
        
        # Record transaction history
        comment = paid_by_combo.get()
        cur.execute("SELECT transaction_id FROM transaction ORDER BY transaction_id DESC LIMIT 1")
        item = cur.fetchone()
        transaction_id = item[0] if item else None
        
        cur.execute("""INSERT INTO transaction_history(learner_id, amount, balance, term_id, comment, transaction_id)
                    VALUES (%s, %s, %s, %s, %s, %s)""",
                    (learner_id, new_amount, new_balance, active, comment, transaction_id))
        my_db.commit()
        
        # Clean up old records if the balance is cleared
        cur.execute("""DELETE FROM transaction_history WHERE learner_id = (
                        SELECT learner_id FROM transaction WHERE balance = 0 AND learner_id = %s 
                        AND term_id = (SELECT term_id FROM term WHERE is_active = 1))""", (learner_id,))
        my_db.commit()
        
        # Clear input and notify success
        amount_paid_entry.delete(0, END)
        disp_label.config(text="Payment saved successfully")
        disp_label.after(4000, lambda: disp_label.config(text=""))
    except Exception as ex:
        messagebox.showerror("Remedial App", f"An error occurred: {ex}")

#retrieving termly pay for terms
#reseting display label
def clear_disp():
    disp_label.after(3000,lambda:disp_label.config(text=""))

#display balances
def display_bal(e):
    clear_tree()
    cur.execute("""SELECT l.grade,l.learner_id,l.first,l.second,
                l.surname,t.amount_paid,t.balance FROM learner l 
                JOIN transaction t  ON l.learner_id=t.learner_id 
                WHERE term_id=(SELECT term_id FROM term WHERE
                is_active=1)
                ORDER BY l.grade DESC""")
    row=cur.fetchall()
    if row:
        for index,row in enumerate(row, start=1):
            fee_tree.insert('',END,values=(index,f"{row[0]}".title(),row[1],f"{row[2]} {row[3]} {row[4]}".title(),row[5],row[6]))
        amount_paid_entry.focus()
        amount_paid_entry.delete(0,END)
    else:
        cur.execute("""SELECT selected_term FROM term WHERE is_active=1""")
        term=cur.fetchone()[0]
        
        messagebox.showinfo("Remedial app",f"No balance records for {term}")
#setting loading default term to combobox
#populating the term combo box
terms=["Term One,2025",
        "Term Two, 2025",
        "Term Three, 2025",
        "Term One,2026",
        "Term Two, 2026",
        "Term Three, 2026",
        "Term One, 2027",
        "Term Two, 2027",
        "Term Three, 2027",
        "Term One, 2028",
        "Term Two, 2028",
        "Term Three, 2028",
        "Term One, 2029",
        "Term Two, 2029",
        "Term Three, 2029",
        "Term One, 2030",
        "Term Two, 2030",
        "Term Three,2030"]

# set selected term as active
def set_default_term(e):
    selected=term_combo.get()
    #deactivating the rest of terms
    cur.execute("UPDATE term SET  is_active=0 WHERE is_active=1")
    my_db.commit()
    #updating selected term to active
    cur.execute("UPDATE term SET is_active=1  WHERE selected_term=%s",(selected,))
    my_db.commit()
    clear_boxes()
    # clear_tree()
    disp_all_learners()
    week_one()
def retrieve_term():
#     #retrieving selected term and setting it to default term
    cur.execute("SELECT selected_term FROM term WHERE is_active=1")
    default=cur.fetchall()
    if default:
        for i in default[0]:
            term_combo.set(i)
    else:
        term_combo.set(terms[0] if terms else '')
              
def display_tr(e=None):
    clear_tree()
    cur.execute("SELECT * FROM teacher")
    row=cur.fetchall()
    if row:
        for index,row in enumerate(row,start=1):
            fee_tree.insert('',END,values=(index,"x",row[0],f'{row[1]} {row[2]} {row[3]}'.title()))
    else:
        messagebox.showinfo("Remedial App","Teachers list is empty")
    clear_boxes()
def learners_displaying(e):
    clear_boxes()
    fee_tree.delete(*fee_tree.get_children())
    #displaying all learners
    if disp_combo.get()==grade_disp[4]:
        cur.execute("SELECT * FROM learner ORDER BY learner_id")
        row=cur.fetchall()
        if row:
            for index,row in enumerate(row,start=1):
                fee_tree.insert('',END,values=(index,f"{row[4]}".title(),row[0],f"{row[1]} {row[2]} {row[3]}".title()))
        else:
            messagebox.showinfo("Remedial App","Learners List is empty")


        return
    if disp_combo.get()==grade_disp[0]:
         messagebox.showwarning("Remedial App","Select Grade to display")
         return
    disp_class=disp_combo.get()
    if disp_combo.get()==grade_disp[1]:
        disp_class="seven"
    elif disp_combo.get()==grade_disp[2]:
        disp_class="eight"
    else:
        disp_class="nine"
    sql="SELECT * FROM learner WHERE grade=%s"
    cur.execute(sql,(disp_class,))
    row=cur.fetchall()
    if row:
        for index,row in enumerate(row,start=1):
            fee_tree.insert('',END,values=(index,f"{row[4]}".title(),row[0],f"{row[1]} {row[2]} {row[3]}".title()))
    else:
        messagebox.showinfo("Remedial App","Learners List is empty")

#clearing entryboxes
def clear_boxes():
        adm_entry.delete(0,END)
        first_entry.delete(0,END)
        second_entry.delete(0,END)
        surname_entry.delete(0,END)
        # grade_combo.set("")

#popup func
def binding(e):
    learner_menu.tk_popup(e.x,e.y)


#deleting learners
def delete_transaction():
        adm=[]
        tuples=[]
        pos=fee_tree.selection()
        if not pos:
            messagebox.showwarning("Remedial App","Select records before attempting to delete")
        else:
            root.bell()
            resp=messagebox.askyesno("Remedial App","Are sure you want to delete\nSelected PAYMENT record(s)\nPermanently?")
            if resp==1:
                for items in pos:
                    values=fee_tree.item(items,"values") 
                    adm.append(values[2])
                    #deleting from fee_tree
                    fee_tree.delete(items)
                for a in adm:
                    tuples.append((a,))
                cur.executemany("""DELETE FROM transaction WHERE learner_id=%s AND
                                term_id=(SELECT term_id FROM term WHERE is_active=1)""",tuples)
                my_db.commit()
            else:
                pass
        clear_boxes()
        

#deleting learners
def delete_learner():
        adm=[]
        tuples=[]
        pos=fee_tree.selection()
        if not pos:
            messagebox.showwarning("Remedial App","Select records before attempting to delete")
        else:
            root.bell()
            resp=messagebox.askyesno("Remedial App","Are sure you want to delete\n LEARNER(S) permanently?")
            if resp==1:
                for items in pos:
                    values=fee_tree.item(items,"values") 
                    adm.append(values[2])
                    #deleting from fee_tree
                    fee_tree.delete(items)
                for a in adm:
                    tuples.append((a,))
                cur.executemany("DELETE FROM learner WHERE learner_id=%s",tuples)
                my_db.commit()
            else:
                pass
        clear_boxes()
        
#deleting teachers
def delete_teacher():
    adm=[]
    tuples=[]
    pos=fee_tree.selection()
    if not pos:
        messagebox.showwarning("Remedial App","Select records before attempting to delete")
    else:
        root.bell()
        resp=messagebox.askyesno("Remedial App","Are sure you want to delete\nTEACHER(S) permanently?")
        if resp==1:
            for items in pos:
                values=fee_tree.item(items,"values") 
                adm.append(values[2])
                fee_tree.delete(items)
                for a in adm:
                    tuples.append((a,))
                cur.executemany("DELETE FROM teacher WHERE teacher_id=%s",tuples)
                my_db.commit()
            else:
                pass
    clear_boxes()
#clearing the treeview
def clear_tree():
    fee_tree.delete(*fee_tree.get_children()) 


#teachers
# #placer
# def placer(e):
#     cord=f"{e.x} x {e.y}"
#     disp_label.config(text=cord)
#     disp_label.after(3000,lambda:disp_label.config(text=""))
    
   
#placing label    
disp_label=Label(root,text="")
disp_label.place(x=615,y=550)
#displal label2

#title Label
igamba=Label(root,text="REMEDIAL SYSTEM",
             font="Helvetica 14 bold")
# igamba.place(x=40,y=0)
    
#defining widgets  
#LEARNER REGISTRATION WIDGETS
learner_regist_frame=LabelFrame(root,text="Teachers/Learners Registration",
                                  font="Helvetica 14 bold",padx=10)
#entry frame
learner_frame_entry=Frame(learner_regist_frame,border=0)
adm_label=Label(learner_frame_entry,text="ADM NO/TR.NO",font="Times 11")
adm_entry=Entry(learner_frame_entry,font="Helvetica")
grade_combo_label=Label(learner_frame_entry,text="GRADE",font="Times 11")
grade=["Seven","Eight","Nine"]
grade_combo=ttk.Combobox(learner_frame_entry,font="Times 12",values=grade,width=20,state="readonly")
# grade_combo.current(0)
first_label=Label(learner_frame_entry,text="FIRST NAME",font="Times 11")
first_entry=Entry(learner_frame_entry,font="Helvetica")
second_label=Label(learner_frame_entry,text="SECOND NAME",font="Times 11")
second_entry=Entry(learner_frame_entry,font="Helvetica")
surname_label=Label(learner_frame_entry,text="SURNAME",font="Times 11")
surname_entry=Entry(learner_frame_entry,font="Helvetica")
#buttons frame
button_frame=Frame(learner_regist_frame)
add_button=Button(button_frame,text="Add/Edit Learner Details",command=lambda:add_learner(None))
delete_button=Button(button_frame,text="Delete Learner",command=delete_learner)


add_button2=Button(button_frame,text="Add Teacher",command=lambda:add_teachers(None))
delete_button2=Button(button_frame,text="Delete Teacher",command=delete_teacher,bg="grey")
#FEE COLLECTION FRAME(widgets definition)
fee_collection_frame=LabelFrame(root,text="Fee Collection",
                                  font="Helvetica 14 bold",padx=10,pady=10)

        
#SET AMOUNT
learner_pay_button=Button(root,text="Set term fee",command=learner_term_pay)
learner_pay_button.place(x=273,y=22)
lnr_amount_label=Label(root,text="Amount",font="Helvetica")
lnr_amount_entry=Entry(root,font="Helvetica 14",width=5)
Grade_label=Label(root,text="Grade",font="Helvetica")
grade_entry_combo=ttk.Combobox(root,values=[7,8,9],font="Helvetica 14",width=3,state="readonly")
# grade9_label=Label(root,text="Grade 9")
# grade9_entry=Entry(root,font="Helvetica 14",width=5)
l_pay_submit_button=Button(root,text="Submit",command=submit_learner_fee)
#setting teacher pay
tr_lesson_pay_button=Button(root,text="Teacher Pay",command=tr_lsn_pay)
tr_lesson_pay_button.place(x=386,y=22)
wk_label=Label(root,text="Weekday",font="Helvetica")
wk_entry=Entry(root,font="Helvetica 15",width=5)
wkend_label=Label(root,text="Weekend",font="Helvetica")
wkend_entry=Entry(root,font="Helvetica 15",width=5)
tr_pay_submit_button=Button(root,text="Submit",background="grey",command=submit_tr_lsn_pay)


term_label=Label(root,text="Term",font="helvetica 14")
term_combo=ttk.Combobox(root,values=terms,font="times 14",width=15,state="readonly")
term_label.place(x=36,y=20)
term_combo.place(x=88,y=22)

#Search learner by id
search_button=Button(root,text="Search",command=search_by_id)
search_button.place(x=900,y=20)
search_entry=Entry(root,font="times 14",width=6)
search_entry.insert(0,"Adm")
search_entry.place(x=959,y=20)
search_label=Label(root)
search_label.place(x=1020,y=20)
week_label=Label(root,text="WEEK",font="times 12")
weeks=["One","Two","Three","Four","Five","Six","Seven",
       "Eight","Nine","Ten","Eleven","Twelve","Thirteen",
       "Fourteen"]
week_combo=ttk.Combobox(root,values=weeks,width=9,font="helvetica 12",state="readonly")
week_combo.place(x=1171,y=20)
week_label.place(x=1116,y=20)
#PROMOTE LEARNERS


#display selected grades
grade_disp=["Display Learners by Grade",
            "Grade Seven",
            "Grade Eight",
            "Grade Nine",
            "All Learners"]
#display selected grades
grade_bal=["Display balances by Grade",
            "Grade Seven",
            "Grade Eight",
            "Grade Nine"]

disp_combo=ttk.Combobox(root,font="Times 12",values=grade_disp,width=25
                        ,state="readonly")
disp_combo.current(0)
disp_combo.place(x=37,y=403)
disp_combo.bind("<<ComboboxSelected>>",learners_displaying)

disp_bal_combo=ttk.Combobox(root,font="Times 12",values=grade_bal,width=25,state="readonly")
disp_bal_combo.current(0)
disp_bal_combo.place(x=37,y=437)
disp_bal_combo.bind("<<ComboboxSelected>>",grade_balance)

tr_disp_button=Button(root,text="Display Teachers",command=display_tr,bg="grey")
tr_disp_button.place(x=268,y=403)
#displaying widgets
#LEARNER REGISTRATION WIDGETS FRAME
learner_regist_frame.place(x=36,y=85)
learner_frame_entry.grid(row=0,column=0,padx=10)
adm_label.grid(row=0,column=0,padx=10,pady=10,sticky=W)
adm_entry.grid(row=0,column=1,pady=5,ipady=3)
grade_combo_label.grid(row=1,column=0,padx=10,pady=10,sticky=W)
grade_combo.grid(row=1,column=1,pady=5,ipady=3)
first_label.grid(row=2,column=0,padx=10,pady=10,sticky=W)
first_entry.grid(row=2,column=1,pady=5,ipady=3)
second_label.grid(row=3,column=0,padx=10,pady=10,sticky=W)
second_entry.grid(row=3,column=1,pady=5,ipady=3)
surname_label.grid(row=4,column=0,padx=10,pady=10,sticky=W)
surname_entry.grid(row=4,column=1,pady=5,ipady=3)
#buttons frame
button_frame.grid(row=5,column=0,padx=10)
add_button.grid(row=0,column=1,padx=10)
delete_button.grid(row=0,column=2)
#TEACHER REGISTRATION FRAME (displaying widgets)
add_button2.grid(row=1,column=1,pady=7,padx=10)
delete_button2.grid(row=1,column=2,pady=7)
#FEE COLLECTION FRAME(widgets display)
fee_collection_frame.place(x=409,y=85)


#style 
style=ttk.Style()
style.configure('Treeview',
               rowheight=25,
               font="times 14")
style.configure("Treeview.Heading",foreground="blue",font="Times 12 ")
style.map('Treeview',background=[("selected","blue")])
#fee collection treeview
fee_tree_frame=Frame(root)
fee_tree_scroll=Scrollbar(fee_tree_frame,orient=VERTICAL)
fee_tree_scroll.pack(side=RIGHT,fill=Y)
fee_tree_frame.place(x=411,y=96)
fee_tree=ttk.Treeview(fee_tree_frame,yscrollcommand=fee_tree_scroll.set,height=10)
fee_tree_scroll.configure(command=fee_tree.yview)
#defining columns
fee_tree["columns"]=('s/no','grade','adm','name','paid','balance')
fee_tree.column("#0",width=0,stretch=NO)
fee_tree.column("s/no",width=40,anchor="center",minwidth=35)
fee_tree.column("grade",width=65,anchor="center",minwidth=50)
fee_tree.column("adm",width=100,minwidth=80,anchor=CENTER)
fee_tree.column("name",width=270,minwidth=180,anchor=W)
fee_tree.column("paid",width=80,minwidth=75,anchor=CENTER)
fee_tree.column("balance",width=80,minwidth=75,anchor=CENTER)
#headings
fee_tree.heading("#0",text="")
fee_tree.heading("s/no",text="#",anchor=CENTER)
fee_tree.heading("grade",text="GRADE",anchor=CENTER)
fee_tree.heading("adm",text="ADM/P.NO",anchor=CENTER)
fee_tree.heading("name",text="NAME",anchor=CENTER)
fee_tree.heading("paid",text="PAID",anchor=CENTER)
fee_tree.heading("balance",text="BAL",anchor=CENTER)
#fee collection frame display
fee_tree.pack(fill=BOTH,expand=True)
root.update_idletasks()
learner_menu=Menu(root,tearoff=0)
learner_menu=Menu(learner_menu,tearoff=0)
learner_menu.add_cascade(label="Clear Entry fields",command=clear_boxes)
learner_menu.add_separator()
learner_menu.add_command(label="Display Teachers",command=lambda:display_tr(None),accelerator="Ctrl+T")
learner_menu.add_separator()
learner_menu.add_command(label="Display fee Balances",command=lambda:display_bal(None),accelerator="Ctrl+B")
learner_menu.add_separator()
learner_menu.add_command(label="Display Learners",command=lambda:disp_all_learners(None),accelerator="Ctrl+L")
learner_menu.add_separator()
learner_menu.add_command(label="Display Transaction history",command=lambda:top_window_func(None),accelerator="Ctrl+H")
learner_menu.add_separator()
learner_menu.add_command(label="Display Teacher Attendance",command=lambda:tr_attendance_func(None),accelerator="Ctrl+A")
learner_menu.add_separator()
learner_menu.add_command(label="Display Teacher Total",command=display_teacher_total)
learner_menu.add_separator()
delete_menu=Menu(learner_menu,tearoff=0)
learner_menu.add_cascade(label="Delete",menu=delete_menu)
delete_menu.add_command(label="Delete Learner",command=delete_learner)
delete_menu.add_separator()
delete_menu.add_command(label="Delete Teacher",command=delete_teacher)
delete_menu.add_separator()
delete_menu.add_command(label="Delete Transaction",command=delete_transaction)
delete_menu.add_separator()
delete_menu.add_command(label="Delete Attendance Record",command=delete_attend_record)

#menu bar
manage_app=Menu(root,tearoff=0)
root.config(menu=manage_app)
file_menu=Menu(manage_app,tearoff=0)
manage_app.add_cascade(menu=file_menu,label="File")
# file_menu.add_command(label="Display transaction history",command=next_win)
file_menu.add_command(label="Display Archived Learners",command=display_archives)
file_menu.add_separator()
file_menu.add_command(label="Teacher Attendance Archives",command=lambda:tr_attendance_archive_func(None),accelerator="Ctrl+C")
file_menu.add_separator()
file_menu.add_command(label="Generate Remedial files",command=lambda:remedial_records_to_excel(None),accelerator="Ctrl+G")
file_menu.add_separator()
file_menu.add_command(label="Generate Class Lists",command= generate_class_list)

promote_menu=Menu(manage_app,tearoff=0)
manage_app.add_cascade(menu=promote_menu,label="Move Learners")
promote_menu.add_command(label="G7 to G8",command=seven_eight)
promote_menu.add_command(label="G8 to G9",command=eight_nine)
promote_menu.add_command(label="G9 to Archive",command=nine_archive)

#TRANSACTIONS
learner_payment_frame=LabelFrame(root,text="Fees Payment",font="times 12 bold"
                                 ,padx=10,pady=10)
learner_payment_frame.place(x=411,y=375)
amount_paid_label=Label(learner_payment_frame,text="AMOUNT PAID",font="times 11")
amount_paid_label.grid(row=0,column=0,padx=(0,5))
amount_paid_entry=Entry(learner_payment_frame,font="Helvetica 14",width=8)
amount_paid_entry.grid(row=0,column=1)
#paid by
paid_by_label=Label(learner_payment_frame,text="RECEIVED FROM",font="times 11")
value=['Learner','Mobile Money','HOI','Another Teacher']
paid_by_combo=ttk.Combobox(learner_payment_frame,values=value,font="times",width=13,state="readonly")
paid_by_combo.current(0)
paid_by_label.grid(row=1,column=0,padx=(0,5),pady=(20))
paid_by_combo.grid(row=1,column=1,pady=(20))
Submit_button=Button(learner_payment_frame,text="Save Changes",command=lambda:make_payment(None))
Submit_button.grid(row=2,column=0)

#TRACKING TEACHERS ATTENDANCE
tr_attend_frame=LabelFrame(root,text="Teacher Attendance",pady=10,padx=10,
                           font="times 12 bold")
tr_attend_frame.place(x=700,y=375)
tr_attend_label=Label(tr_attend_frame,text="SESSION",font="times 11")
tr_attend_label.grid(row=0,column=0,sticky=W)
tr_attend=["Morning","Evening","Saturday"]
tr_attend_combo=ttk.Combobox(tr_attend_frame,values=tr_attend,font="times",width=7,state="readonly")
tr_attend_combo.grid(row=0,column=1,padx=(5))
track_grade=["Seven","Eight","Nine"]
track_grade_label=Label(tr_attend_frame,text="GRADE",font="times 11")
track_grade_label.grid(row=0,column=2,padx=(5,5))
track_grade_combo=ttk.Combobox(tr_attend_frame,font="times",values=
                               track_grade,width=5,state="readonly")
track_grade_combo.grid(row=0,column=3)
subject=["Maths","Eng","Kisw","INT","SS","AGN","PTC","CRA","CRE","PPI"]
subject_label=Label(tr_attend_frame,text="SUBJECT",font="times 11")
subject_label.grid(row=1,column=2,padx=(5,5))
subject_combo=ttk.Combobox(tr_attend_frame,font="times",values=
                               subject,width=5,state="readonly")
subject_combo.grid(row=1,column=3)
token_label=Label(tr_attend_frame,text="TOKEN",font="times 11")
token_label.grid(row=1,column=0,pady=20,sticky=W)
token_entry=Entry(tr_attend_frame,font="Helvetica 14",width=6)
token_entry.grid(row=1,column=1,padx=5)
capture_button=Button(tr_attend_frame,text="Save Attnd",command=track_teacher_attendance)
capture_button.grid(row=2,column=0,padx=10)
submit_token=Button(tr_attend_frame,text="Pay Token",command=pay_token)
submit_token.grid(row=2,column=1)
search_entry.bind("<FocusIn>",lambda e:search_entry.delete(0,END))
#date
# today_date=date.today()
# d=today_date.strftime("%d")
# m=today_date.strftime("%m")
# y=today_date.strftime("%Y")
# date_label=Label(root,font=("Helvetica 12"))
# date_label.place(x=1173,y=60)
# date_label.config(text=f"{d}-{m}-{y}")
#separator
sep=ttk.Separator(root,orient=VERTICAL)
# sep.place(x=711,y=429,height=150)
# 411,447

# #event bindings
# root.bind("<Button-1>",placer)

fee_tree.bind("<ButtonRelease>",populate)
root.bind("<Button-3>",binding)
# term_combo.bind("<<ComboboxSelected>>", set_term)
fee_tree_scroll.bind("<MouseWheel>")
#binding the update button
root.bind("<Return>",make_payment)
root.bind("<Control-b>",display_bal)
root.bind("<Control-B>",display_bal)
root.bind("<Control-l>",disp_all_learners)
root.bind("<Control-L>",disp_all_learners)
root.bind("<Control-t>",display_tr)
root.bind("<Control-T>",display_tr)
root.bind("<Control-h>",top_window_func)
root.bind("<Control-H>",top_window_func)
root.bind("<Control-a>",tr_attendance_func)
root.bind("<Control-A>",tr_attendance_func)
root.bind("<Control-g>",remedial_records_to_excel)
root.bind("<Control-G>",remedial_records_to_excel)
root.bind("<Control-C>",tr_attendance_archive_func)
root.bind("<Control-c>",tr_attendance_archive_func)
#saving learners and teachers records
root.bind("<Control-z>",add_learner)
root.bind("<Control-Z>",add_learner)
root.bind("<Control-x>",add_teachers)
root.bind("<Control-X>",add_teachers)
#closing app
def confirm_close():
    root.bell()
    res=messagebox.askyesno("Remedial App","Are you sure you want to close\nRemedial App?")
    if res:
        root.destroy()
      


root.protocol("WM_DELETE_WINDOW",confirm_close)

term_combo.bind("<<ComboboxSelected>>",set_default_term)
week_combo.bind("<<ComboboxSelected>>",set_default_week)
retrieve_term()
retrieve_week()
disp_all_learners()

root.update()
root.mainloop()
my_db.close()
cur.close()






